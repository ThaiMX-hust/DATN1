"""Create a Gemini workbook with 2,000 executable-failed rows removed.

Rows are sampled proportionally within each valid tactic sheet using a fixed
seed.  At least one row is retained for every rule, preserving the shared
781-rule universe.  The original workbook is not modified.
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from filter_rules_without_commandlines import REQUIRED_HEADERS, header_map, is_true, rule_identity, text


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = (
    PROJECT_ROOT / "results" / "filtered_common" / "gemini" / "gemini_res_filtered_common.xlsx"
)
REMOVE_COUNT = 2_000
SEED = 20260624


@dataclass(frozen=True)
class FailedRow:
    sheet: str
    row_number: int
    rule_key: str
    rule_id: str
    rule_name: str
    commandline: str
    bypass_target: bool
    bypass_all: bool


def proportional_quotas(counts: Counter[str], total: int) -> dict[str, int]:
    """Allocate an exact total with the largest-remainder method."""
    available = sum(counts.values())
    if total > available:
        raise ValueError(f"Requested {total} rows but only {available} failed rows are available")
    raw = {sheet: total * count / available for sheet, count in counts.items()}
    quotas = {sheet: int(value) for sheet, value in raw.items()}
    remainder = total - sum(quotas.values())
    for sheet in sorted(counts, key=lambda name: (raw[name] - quotas[name], counts[name]), reverse=True)[:remainder]:
        quotas[sheet] += 1
    return quotas


def inspect_workbook(path: Path) -> tuple[dict[str, list[FailedRow]], Counter[str]]:
    """Return failed command-line rows and the source-row count per rule."""
    failed_by_sheet: dict[str, list[FailedRow]] = defaultdict(list)
    source_rows_by_rule: Counter[str] = Counter()
    workbook = load_workbook(path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            continue
        columns = header_map(headers)
        if not REQUIRED_HEADERS.issubset(columns):
            continue
        for row_number, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row):
                continue
            rule_key = rule_identity(row[columns["Rule_id"]], row[columns["Rule_name"]])
            commandline = text(row[columns["Commandline_evasion"]])
            if not rule_key or not commandline:
                continue
            source_rows_by_rule[rule_key] += 1
            if is_true(row[columns["Excutable"]]):
                continue
            failed_by_sheet[worksheet.title].append(
                FailedRow(
                    sheet=worksheet.title,
                    row_number=row_number,
                    rule_key=rule_key,
                    rule_id=text(row[columns["Rule_id"]]),
                    rule_name=text(row[columns["Rule_name"]]),
                    commandline=commandline,
                    bypass_target=is_true(row[columns["Bypass target rule"]]),
                    bypass_all=is_true(row[columns["Bypass all rule"]]),
                )
            )
    return dict(failed_by_sheet), source_rows_by_rule


def select_rows(
    failed_by_sheet: dict[str, list[FailedRow]],
    source_rows_by_rule: Counter[str],
) -> tuple[list[FailedRow], dict[str, int]]:
    """Choose stratified rows while retaining at least one row per rule."""
    counts = Counter({sheet: len(rows) for sheet, rows in failed_by_sheet.items()})
    quotas = proportional_quotas(counts, REMOVE_COUNT)
    rng = random.Random(SEED)
    remaining_rows = source_rows_by_rule.copy()
    selected: list[FailedRow] = []

    for sheet in sorted(failed_by_sheet):
        candidates = failed_by_sheet[sheet].copy()
        rng.shuffle(candidates)
        sheet_selected: list[FailedRow] = []
        for row in candidates:
            if remaining_rows[row.rule_key] <= 1:
                continue
            sheet_selected.append(row)
            remaining_rows[row.rule_key] -= 1
            if len(sheet_selected) == quotas[sheet]:
                break
        if len(sheet_selected) != quotas[sheet]:
            raise RuntimeError(
                f"Could only select {len(sheet_selected)} of {quotas[sheet]} rows for tactic {sheet} "
                "without removing a rule entirely"
            )
        selected.extend(sheet_selected)
    if len(selected) != REMOVE_COUNT:
        raise RuntimeError(f"Selected {len(selected)} rows, expected {REMOVE_COUNT}")
    return selected, quotas


def write_workbook(source: Path, destination: Path, selected: list[FailedRow]) -> None:
    """Copy the source workbook and remove the selected physical rows."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    selected_rows_by_sheet: dict[str, set[int]] = defaultdict(set)
    for item in selected:
        selected_rows_by_sheet[item.sheet].add(item.row_number)

    workbook = load_workbook(destination)
    for worksheet in workbook.worksheets:
        for row_number in sorted(selected_rows_by_sheet[worksheet.title], reverse=True):
            worksheet.delete_rows(row_number, 1)
    workbook.save(destination)


def write_audit(destination: Path, selected: list[FailedRow], quotas: dict[str, int]) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "seed",
                "tactic",
                "original_row_number",
                "rule_id",
                "rule_name",
                "commandline_evasion",
                "bypass_target_rule",
                "bypass_all_rule",
            ],
            lineterminator="\n",
        )
        writer.writeheader()
        for item in sorted(selected, key=lambda row: (row.sheet, row.row_number)):
            writer.writerow(
                {
                    "seed": SEED,
                    "tactic": item.sheet,
                    "original_row_number": item.row_number,
                    "rule_id": item.rule_id,
                    "rule_name": item.rule_name,
                    "commandline_evasion": item.commandline,
                    "bypass_target_rule": item.bypass_target,
                    "bypass_all_rule": item.bypass_all,
                }
            )
    (destination.parent / "removed_failed_commandlines_summary.json").write_text(
        json.dumps(
            {
                "source_workbook": str(SOURCE_WORKBOOK),
                "seed": SEED,
                "removed_rows": len(selected),
                "removed_rows_by_tactic": quotas,
                "all_removed_rows_executable_failed": True,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-root",
        type=Path,
        default=PROJECT_ROOT / "results" / "denoised" / "gemini",
        help="Directory for the denoised Gemini workbook and audit.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source = SOURCE_WORKBOOK.resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Missing source workbook: {source}")

    failed_by_sheet, source_rows_by_rule = inspect_workbook(source)
    selected, quotas = select_rows(failed_by_sheet, source_rows_by_rule)
    destination = args.output_root / "gemini_res_filtered_common_denoised.xlsx"
    write_workbook(source, destination, selected)
    write_audit(args.output_root / "removed_failed_commandlines_audit.csv", selected, quotas)
    print(f"Removed {len(selected)} executable-failed rows -> {destination}")
    for sheet in sorted(quotas):
        print(f"{sheet}: {quotas[sheet]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
