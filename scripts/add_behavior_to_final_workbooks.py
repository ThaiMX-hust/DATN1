"""Add the manually verified behavior label to final result workbooks."""

from __future__ import annotations

import argparse
import csv
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EVASION_CSV = PROJECT_ROOT / "results" / "results_final" / "evasion_final.csv"
DEFAULT_RESULTS_DIR = PROJECT_ROOT / "results" / "results_final"
SUMMARY_SHEETS = {"fill_summary", "summary", "thong_ke", "thống_kê"}
TACTIC_REQUIRED_HEADERS = {"Rule_name", "Commandline_evasion", "Bypass all rule"}
BEHAVIOR_HEADER = "behavior"
TRUE_VALUES = {"true", "1", "1.0", "yes", "y"}


def cell_text(value: Any) -> str:
    """Return a stable text value for keys."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace("\r\n", "\n")


def is_true(value: Any) -> bool:
    """Return whether a cell value is true."""
    return value is True or cell_text(value).lower() in TRUE_VALUES


def normalize_sheet_name(value: Any) -> str:
    """Normalize worksheet names for skip checks."""
    return cell_text(value).lower().replace(" ", "_")


def header_map(worksheet: Any) -> dict[str, int]:
    """Return header text to one-based column index."""
    return {
        cell_text(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
        if cell_text(worksheet.cell(1, column).value)
    }


def read_behavior_keys(evasion_csv: Path) -> set[tuple[str, str]]:
    """Read manually verified evasion keys from evasion_final.csv."""
    with evasion_csv.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        if not reader.fieldnames:
            raise ValueError(f"empty CSV: {evasion_csv}")
        missing = sorted({"rule_name", "commandline_evasion"} - set(reader.fieldnames))
        if missing:
            raise ValueError(f"{evasion_csv} missing column(s): {', '.join(missing)}")

        keys = {
            (cell_text(row["rule_name"]), cell_text(row["commandline_evasion"]))
            for row in reader
            if cell_text(row["rule_name"]) and cell_text(row["commandline_evasion"])
        }
    return keys


def default_workbooks() -> list[Path]:
    """Return the three final result workbooks under results/results_final."""
    return sorted(DEFAULT_RESULTS_DIR.glob("*/*_res_final.xlsx"))


def is_tactic_sheet(sheet_name: str, headers: dict[str, int]) -> bool:
    """Return whether the worksheet has the final result tactic schema."""
    if normalize_sheet_name(sheet_name) in SUMMARY_SHEETS:
        return False
    return TACTIC_REQUIRED_HEADERS.issubset(headers)


def ensure_behavior_column(worksheet: Any, headers: dict[str, int]) -> int:
    """Return the behavior column, appending it if needed."""
    existing = headers.get(BEHAVIOR_HEADER)
    if existing:
        return existing

    behavior_column = worksheet.max_column + 1
    worksheet.cell(1, behavior_column, BEHAVIOR_HEADER)

    source_column = headers.get("Bypass all rule") or headers.get("Bypass target rule")
    if source_column:
        source_header = worksheet.cell(1, source_column)
        behavior_header = worksheet.cell(1, behavior_column)
        behavior_header._style = copy(source_header._style)
        behavior_header.font = copy(source_header.font)
        behavior_header.fill = copy(source_header.fill)
        behavior_header.border = copy(source_header.border)
        behavior_header.alignment = copy(source_header.alignment)
        behavior_header.number_format = source_header.number_format

    worksheet.column_dimensions[get_column_letter(behavior_column)].width = 13
    return behavior_column


def copy_row_style(worksheet: Any, source_column: int | None, behavior_column: int, row_number: int) -> None:
    """Copy the adjacent boolean-column style to the behavior cell."""
    if not source_column:
        return
    source_cell = worksheet.cell(row_number, source_column)
    target_cell = worksheet.cell(row_number, behavior_column)
    target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format


def extend_ranges(worksheet: Any, last_row: int, last_col: int) -> int:
    """Extend filters/tables so the new behavior column is included."""
    worksheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    updated_tables = 0
    for table_name in list(worksheet.tables):
        table = worksheet.tables[table_name]
        min_col, min_row, table_max_col, table_max_row = range_boundaries(table.ref)
        new_max_col = max(table_max_col, last_col)
        new_max_row = max(table_max_row, last_row)
        new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(new_max_col)}{new_max_row}"
        if new_ref == table.ref:
            continue
        table.ref = new_ref
        if table.autoFilter is not None:
            table.autoFilter.ref = new_ref
        updated_tables += 1
    return updated_tables


def update_workbook(workbook_path: Path, behavior_keys: set[tuple[str, str]]) -> dict[str, int]:
    """Update one workbook and return update counts."""
    workbook = load_workbook(workbook_path)
    tactic_sheets = 0
    source_rows = 0
    behavior_true = 0
    behavior_false = 0
    evasion_key_without_bypass_all = 0
    table_updates = 0

    for worksheet in workbook.worksheets:
        headers = header_map(worksheet)
        if not is_tactic_sheet(worksheet.title, headers):
            continue

        tactic_sheets += 1
        behavior_column = ensure_behavior_column(worksheet, headers)
        source_style_column = headers.get("Bypass all rule") or headers.get("Bypass target rule")
        command_column = headers.get("Commandline_evasion") or headers.get("Command_match_rule")
        rule_column = headers["Rule_name"]
        bypass_all_column = headers["Bypass all rule"]

        for row_number in range(2, worksheet.max_row + 1):
            if not any(
                worksheet.cell(row_number, column).value not in (None, "")
                for column in range(1, worksheet.max_column + 1)
                if column != behavior_column
            ):
                continue

            source_rows += 1
            rule_name = cell_text(worksheet.cell(row_number, rule_column).value)
            commandline = cell_text(worksheet.cell(row_number, command_column).value) if command_column else ""
            in_evasion_final = bool(rule_name and commandline and (rule_name, commandline) in behavior_keys)
            bypass_all = is_true(worksheet.cell(row_number, bypass_all_column).value)
            value = in_evasion_final and bypass_all
            worksheet.cell(row_number, behavior_column, value)
            copy_row_style(worksheet, source_style_column, behavior_column, row_number)
            if value:
                behavior_true += 1
            else:
                behavior_false += 1
                if in_evasion_final and not bypass_all:
                    evasion_key_without_bypass_all += 1

        table_updates += extend_ranges(worksheet, worksheet.max_row, max(worksheet.max_column, behavior_column))

    workbook.save(workbook_path)
    return {
        "tactic_sheets": tactic_sheets,
        "source_rows": source_rows,
        "behavior_true": behavior_true,
        "behavior_false": behavior_false,
        "evasion_key_without_bypass_all": evasion_key_without_bypass_all,
        "table_updates": table_updates,
    }


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Add/update behavior=true labels in final result workbooks from evasion_final.csv."
    )
    parser.add_argument("--evasion-csv", type=Path, default=DEFAULT_EVASION_CSV)
    parser.add_argument(
        "--workbook",
        action="append",
        dest="workbooks",
        type=Path,
        help="Workbook to update. Repeat for multiple files. Defaults to all *_res_final.xlsx files.",
    )
    return parser.parse_args()


def main() -> int:
    """Run the behavior label update."""
    args = parse_args()
    evasion_csv = args.evasion_csv.resolve()
    workbooks = [path.resolve() for path in args.workbooks] if args.workbooks else default_workbooks()

    if not evasion_csv.is_file():
        raise FileNotFoundError(f"evasion CSV not found: {evasion_csv}")
    if not workbooks:
        raise FileNotFoundError(f"no *_res_final.xlsx workbooks found under {DEFAULT_RESULTS_DIR}")

    behavior_keys = read_behavior_keys(evasion_csv)
    print(f"Loaded {len(behavior_keys)} behavior keys from {evasion_csv}")

    for workbook_path in workbooks:
        if not workbook_path.is_file():
            raise FileNotFoundError(f"workbook not found: {workbook_path}")
        counts = update_workbook(workbook_path, behavior_keys)
        print(
            f"{workbook_path}: sheets={counts['tactic_sheets']}, rows={counts['source_rows']}, "
            f"behavior_true={counts['behavior_true']}, behavior_false={counts['behavior_false']}, "
            f"evasion_key_without_bypass_all={counts['evasion_key_without_bypass_all']}"
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
