"""Summarize command-line and rule results from final_result.xlsx."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = PROJECT_ROOT / "results" / "qwen32b" / "qwen32b_res.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT /  "final_result_stats"/ "qwen32b"

SUMMARY_SHEETS = {"fill_summary", "summary"}
TRUE_VALUES = {"true", "1", "1.0", "yes", "y"}
FALSE_VALUES = {"false", "0", "0.0", "no", "n"}

SUMMARY_FIELDNAMES = [
    "scope",
    "tactic",
    "source_rows",
    "commandline_total",
    "commandline_executable_success",
    "commandline_bypass_target_rule",
    "commandline_bypass_all_rule",
    "rule_total",
    "rule_with_commandline",
    "rule_with_executable_success",
    "rule_with_bypass_target_rule",
    "rule_with_bypass_all_rule",
]

COMMANDLINE_FIELDNAMES = [
    "scope",
    "tactic",
    "commandline_total",
    "commandline_executable_success",
    "commandline_bypass_target_rule",
    "commandline_bypass_all_rule",
]

RULE_FIELDNAMES = [
    "scope",
    "tactic",
    "rule_total",
    "rule_with_commandline",
    "rule_with_executable_success",
    "rule_with_bypass_target_rule",
    "rule_with_bypass_all_rule",
]


@dataclass(frozen=True)
class ResultRecord:
    """One workbook row that contains a command line."""

    sheet: str
    rule_name: str
    rule_id: str
    commandline: str
    llm: str
    executable: bool
    bypass_target: bool
    bypass_all: bool

    @property
    def rule_key(self) -> str:
        """Return the rule identity used for rule-level statistics."""
        return self.rule_name

    @property
    def commandline_key(self) -> tuple[str, str, str, str]:
        """Return the commandline identity used for unique-mode statistics."""
        return (self.rule_name, self.rule_id, self.commandline, self.llm)

    @property
    def sheet_commandline_key(self) -> tuple[str, str, str, str, str]:
        """Return a sheet-aware commandline identity for row-mode overall stats."""
        return (self.sheet, self.rule_name, self.rule_id, self.commandline, self.llm)


def normalize_sheet_name(value: Any) -> str:
    """Normalize a sheet name for skip checks."""
    return str(value or "").strip().lower().replace(" ", "_")


def cell_text(value: Any) -> str:
    """Return a stable text value for keys and CSV output."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_bool(value: Any) -> bool:
    """Normalize workbook boolean-ish values."""
    if isinstance(value, bool):
        return value
    text = cell_text(value).lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return False


def header_map(headers: tuple[Any, ...]) -> dict[str, int]:
    """Return header text to zero-based tuple index."""
    return {str(value).strip(): index for index, value in enumerate(headers) if value not in (None, "")}


def first_value(row: tuple[Any, ...], headers: dict[str, int], names: list[str]) -> Any:
    """Return the first non-empty row value for a list of possible headers."""
    for name in names:
        index = headers.get(name)
        if index is None or index >= len(row):
            continue
        value = row[index]
        if value not in (None, ""):
            return value
    return None


def read_workbook_records(workbook_path: Path) -> tuple[list[ResultRecord], dict[str, set[str]], Counter[str], list[dict[str, str]]]:
    """Read tactic sheets from a workbook."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    records: list[ResultRecord] = []
    rules_by_sheet: dict[str, set[str]] = {}
    source_rows_by_sheet: Counter[str] = Counter()
    skipped_sheets: list[dict[str, str]] = []

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        if normalize_sheet_name(sheet_name) in SUMMARY_SHEETS:
            continue

        row_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = header_map(next(row_iter))
        except StopIteration:
            skipped_sheets.append({"sheet": sheet_name, "reason": "empty sheet"})
            continue

        required = {"Rule_name", "Excutable", "Bypass target rule", "Bypass all rule"}
        missing = sorted(required - set(headers))
        if missing:
            skipped_sheets.append({"sheet": sheet_name, "reason": f"missing header(s): {', '.join(missing)}"})
            continue

        sheet_rules: set[str] = set()
        for row in row_iter:
            if not any(value not in (None, "") for value in row):
                continue

            source_rows_by_sheet[sheet_name] += 1
            rule_name = cell_text(first_value(row, headers, ["Rule_name"]))
            if not rule_name:
                continue
            sheet_rules.add(rule_name)

            commandline = cell_text(first_value(row, headers, ["Commandline_evasion", "Command_match_rule"]))
            if not commandline:
                continue

            records.append(
                ResultRecord(
                    sheet=sheet_name,
                    rule_name=rule_name,
                    rule_id=cell_text(first_value(row, headers, ["Rule_id"])),
                    commandline=commandline,
                    llm=cell_text(first_value(row, headers, ["LLM"])),
                    executable=normalize_bool(first_value(row, headers, ["Excutable", "Executable"])),
                    bypass_target=normalize_bool(first_value(row, headers, ["Bypass target rule"])),
                    bypass_all=normalize_bool(first_value(row, headers, ["Bypass all rule"])),
                )
            )

        rules_by_sheet[sheet_name] = sheet_rules

    return records, rules_by_sheet, source_rows_by_sheet, skipped_sheets


def aggregate_records(records: list[ResultRecord], count_mode: str, overall: bool) -> list[ResultRecord]:
    """Aggregate commandline records for unique mode, or return row records."""
    if count_mode == "rows":
        return records

    grouped: dict[tuple[Any, ...], ResultRecord] = {}
    for record in records:
        key = record.commandline_key if overall else record.sheet_commandline_key
        existing = grouped.get(key)
        if existing is None:
            grouped[key] = record
            continue
        grouped[key] = ResultRecord(
            sheet=existing.sheet,
            rule_name=existing.rule_name,
            rule_id=existing.rule_id,
            commandline=existing.commandline,
            llm=existing.llm,
            executable=existing.executable or record.executable,
            bypass_target=existing.bypass_target or record.bypass_target,
            bypass_all=existing.bypass_all or record.bypass_all,
        )
    return list(grouped.values())


def summarize_scope(
    scope: str,
    tactic: str,
    records: list[ResultRecord],
    rule_total: int,
    source_rows: int,
    count_mode: str,
    overall: bool,
) -> dict[str, Any]:
    """Summarize one tactic sheet or the whole workbook."""
    commandlines = aggregate_records(records, count_mode=count_mode, overall=overall)
    rules_with_commandline = {record.rule_key for record in commandlines}
    rules_with_executable = {record.rule_key for record in commandlines if record.executable}
    rules_with_bypass_target = {record.rule_key for record in commandlines if record.bypass_target}
    rules_with_bypass_all = {record.rule_key for record in commandlines if record.bypass_all}

    return {
        "scope": scope,
        "tactic": tactic,
        "source_rows": source_rows,
        "commandline_total": len(commandlines),
        "commandline_executable_success": sum(1 for record in commandlines if record.executable),
        "commandline_bypass_target_rule": sum(1 for record in commandlines if record.bypass_target),
        "commandline_bypass_all_rule": sum(1 for record in commandlines if record.bypass_all),
        "rule_total": rule_total,
        "rule_with_commandline": len(rules_with_commandline),
        "rule_with_executable_success": len(rules_with_executable),
        "rule_with_bypass_target_rule": len(rules_with_bypass_target),
        "rule_with_bypass_all_rule": len(rules_with_bypass_all),
    }


def build_summary(
    records: list[ResultRecord],
    rules_by_sheet: dict[str, set[str]],
    source_rows_by_sheet: Counter[str],
    count_mode: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Build per-tactic and overall summary rows."""
    per_sheet_rows: list[dict[str, Any]] = []
    for sheet_name in rules_by_sheet:
        sheet_records = [record for record in records if record.sheet == sheet_name]
        per_sheet_rows.append(
            summarize_scope(
                scope="tactic",
                tactic=sheet_name,
                records=sheet_records,
                rule_total=len(rules_by_sheet[sheet_name]),
                source_rows=source_rows_by_sheet[sheet_name],
                count_mode=count_mode,
                overall=False,
            )
        )

    all_rules = set().union(*rules_by_sheet.values()) if rules_by_sheet else set()
    overall_row = summarize_scope(
        scope="overall",
        tactic="ALL",
        records=records,
        rule_total=len(all_rules),
        source_rows=sum(source_rows_by_sheet.values()),
        count_mode=count_mode,
        overall=True,
    )
    return per_sheet_rows, overall_row


def commandline_rows(summary_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Extract commandline-level table rows."""
    return [{field: row[field] for field in COMMANDLINE_FIELDNAMES} for row in summary_rows]


def rule_rows(summary_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Extract rule-level table rows."""
    return [{field: row[field] for field in RULE_FIELDNAMES} for row in summary_rows]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    """Write rows to CSV."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def append_table(workbook: Workbook, sheet_name: str, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    """Append a table to an output workbook sheet."""
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(fieldnames)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append([row.get(field) for field in fieldnames])
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        width = max(len(cell_text(cell.value)) for cell in column_cells) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 42)


def write_excel_summary(
    path: Path,
    per_tactic_rows: list[dict[str, Any]],
    overall_row: dict[str, Any],
    skipped_sheets: list[dict[str, str]],
) -> None:
    """Write an XLSX report with separate summary tables."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    all_summary_rows = [overall_row, *per_tactic_rows]

    append_table(workbook, "Overall", SUMMARY_FIELDNAMES, [overall_row])
    append_table(workbook, "By_Tactic", SUMMARY_FIELDNAMES, per_tactic_rows)
    append_table(workbook, "Commandline", COMMANDLINE_FIELDNAMES, commandline_rows(all_summary_rows))
    append_table(workbook, "Rule", RULE_FIELDNAMES, rule_rows(all_summary_rows))
    append_table(workbook, "Skipped_Sheets", ["sheet", "reason"], skipped_sheets)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(
        description="Summarize commandline and rule results from final_result.xlsx by tactic sheet."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--count-mode",
        choices=["unique", "rows"],
        default="unique",
        help=(
            "unique: count unique Rule_name+Rule_id+Commandline+LLM records; "
            "rows: count worksheet row appearances."
        ),
    )
    return parser.parse_args()


def main() -> int:
    """Run the final_result summarizer."""
    args = parse_args()
    workbook_path = args.workbook.resolve()
    output_dir = args.output_dir.resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(f"workbook not found: {workbook_path}")

    records, rules_by_sheet, source_rows_by_sheet, skipped_sheets = read_workbook_records(workbook_path)
    per_tactic_rows, overall_row = build_summary(
        records,
        rules_by_sheet,
        source_rows_by_sheet,
        count_mode=args.count_mode,
    )
    summary_rows = [overall_row, *per_tactic_rows]

    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "by_tactic_summary.csv", SUMMARY_FIELDNAMES, per_tactic_rows)
    write_csv(output_dir / "overall_summary.csv", SUMMARY_FIELDNAMES, [overall_row])
    write_csv(output_dir / "commandline_summary.csv", COMMANDLINE_FIELDNAMES, commandline_rows(summary_rows))
    write_csv(output_dir / "rule_summary.csv", RULE_FIELDNAMES, rule_rows(summary_rows))
    write_excel_summary(output_dir / "final_result_statistics.xlsx", per_tactic_rows, overall_row, skipped_sheets)

    payload = {
        "workbook": str(workbook_path),
        "count_mode": args.count_mode,
        "overall": overall_row,
        "by_tactic": per_tactic_rows,
        "skipped_sheets": skipped_sheets,
    }
    (output_dir / "final_result_statistics.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Read {len(records)} commandline rows from {workbook_path}")
    print(f"Count mode: {args.count_mode}")
    print(f"Tactic sheets: {len(per_tactic_rows)}")
    print(f"Overall commandlines: {overall_row['commandline_total']}")
    print(f"Overall rules with commandline: {overall_row['rule_with_commandline']}")
    print(f"Wrote reports to {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
