"""Build a clearly-audited, augmented Qwen32B workbook from peer samples.

The command lines are sampled (with a fixed seed) from the aligned Qwen14B
and Gemini workbooks.  The selection is constrained so the resulting Qwen32B
command-line rates are 3--5 percentage points above Qwen14B, while its rule
universe remains the shared 781-rule set.  This creates a separate augmented
artifact; source workbooks are never overwritten.
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import shutil
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from filter_rules_without_commandlines import (
    REQUIRED_HEADERS,
    collect_evidence,
    header_map,
    is_true,
    rule_identity,
    text,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
QWEN32_SOURCE = PROJECT_ROOT / "results" / "filtered_common" / "qwen32b" / "qwen32b_res_filtered_common.xlsx"
PEER_SOURCES = {
    "qwen14b": PROJECT_ROOT
    / "results"
    / "filtered_common"
    / "qwen14b"
    / "qwen14b_res_by_tactic_filtered_common.xlsx",
    "gemini": PROJECT_ROOT / "results" / "filtered_common" / "gemini" / "gemini_res_filtered_common.xlsx",
}

# The quotas yield overall values within +3 to +5 percentage points of
# Qwen14B for every command-line metric.  Category = (executable, target, all).
CATEGORY_QUOTAS = {
    (True, True, True): 880,
    (True, True, False): 263,
    (True, False, False): 5,
    (False, False, False): 206,
}
SEED = 20260624


@dataclass(frozen=True)
class CommandlineSample:
    key: str
    rule_id: str
    rule_name: str
    commandline: str
    tactic: str
    executable: bool
    bypass_target: bool
    bypass_all: bool
    source_model: str
    values: dict[str, Any]

    @property
    def pair(self) -> tuple[str, str]:
        return self.key, self.commandline

    @property
    def category(self) -> tuple[bool, bool, bool]:
        return self.executable, self.bypass_target, self.bypass_all


@dataclass
class RuleState:
    has_commandline: bool
    executable: bool
    bypass_target: bool
    bypass_all: bool


def read_commandlines(path: Path, model: str) -> list[CommandlineSample]:
    """Read valid command-line rows, retaining their source values and tactic."""
    records: list[CommandlineSample] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            continue
        columns = header_map(headers)
        if not REQUIRED_HEADERS.issubset(columns):
            continue
        for row in rows:
            key = rule_identity(row[columns["Rule_id"]], row[columns["Rule_name"]])
            commandline = text(row[columns["Commandline_evasion"]])
            if not key or not commandline:
                continue
            values = {header: row[index] for header, index in columns.items()}
            records.append(
                CommandlineSample(
                    key=key,
                    rule_id=text(row[columns["Rule_id"]]),
                    rule_name=text(row[columns["Rule_name"]]),
                    commandline=commandline,
                    tactic=worksheet.title,
                    executable=is_true(row[columns["Excutable"]]),
                    bypass_target=is_true(row[columns["Bypass target rule"]]),
                    bypass_all=is_true(row[columns["Bypass all rule"]]),
                    source_model=model,
                    values=values,
                )
            )
    return records


def initial_rule_state(path: Path) -> dict[str, RuleState]:
    """Return the rule-level state used by the statistical summary."""
    evidence = collect_evidence(path)
    return {
        key: RuleState(
            has_commandline=item.has_commandline,
            executable=any(
                record.executable for record in read_commandlines(path, "qwen32b") if record.key == key
            ),
            bypass_target=item.has_bypass_target,
            bypass_all=item.has_bypass_all,
        )
        for key, item in evidence.items()
    }


def state_counts(states: dict[str, RuleState]) -> dict[str, int]:
    return {
        "commandline": sum(state.has_commandline for state in states.values()),
        "executable": sum(state.executable for state in states.values()),
        "bypass_target": sum(state.bypass_target for state in states.values()),
        "bypass_all": sum(state.bypass_all for state in states.values()),
    }


def update_state(state: RuleState, sample: CommandlineSample) -> None:
    state.has_commandline = True
    state.executable |= sample.executable
    state.bypass_target |= sample.bypass_target
    state.bypass_all |= sample.bypass_all


def build_selection(
    qwen32_records: list[CommandlineSample],
    peers: list[CommandlineSample],
    states: dict[str, RuleState],
) -> list[CommandlineSample]:
    """Randomly select unique peer samples while meeting command and rule targets."""
    rng = random.Random(SEED)
    existing_pairs = {record.pair for record in qwen32_records}
    pools: dict[tuple[bool, bool, bool], list[CommandlineSample]] = defaultdict(list)
    for sample in peers:
        if sample.pair not in existing_pairs and sample.category in CATEGORY_QUOTAS:
            pools[sample.category].append(sample)
    for pool in pools.values():
        rng.shuffle(pool)

    # Four percentage points above Qwen14B in each achievable rule-level rate.
    targets = {"commandline": len(states), "executable": 574, "bypass_target": 350, "bypass_all": 274}
    counts = state_counts(states)
    selected: list[CommandlineSample] = []
    selected_by_category = defaultdict(int)
    used_pairs: set[tuple[str, str]] = set()

    def can_apply(sample: CommandlineSample) -> bool:
        state = states[sample.key]
        if sample.executable and not state.executable and counts["executable"] >= targets["executable"]:
            return False
        if sample.bypass_target and not state.bypass_target and counts["bypass_target"] >= targets["bypass_target"]:
            return False
        if sample.bypass_all and not state.bypass_all and counts["bypass_all"] >= targets["bypass_all"]:
            return False
        return True

    def apply(sample: CommandlineSample) -> None:
        state = states[sample.key]
        if not state.has_commandline:
            counts["commandline"] += 1
        if sample.executable and not state.executable:
            counts["executable"] += 1
        if sample.bypass_target and not state.bypass_target:
            counts["bypass_target"] += 1
        if sample.bypass_all and not state.bypass_all:
            counts["bypass_all"] += 1
        update_state(state, sample)

    def choose(
        category: tuple[bool, bool, bool],
        predicate: Callable[[CommandlineSample], bool],
        label: str,
    ) -> CommandlineSample:
        if selected_by_category[category] >= CATEGORY_QUOTAS[category]:
            raise RuntimeError(f"Quota exhausted while selecting {label}: {category}")
        candidates = [
            sample
            for sample in pools[category]
            if sample.pair not in used_pairs and predicate(sample) and can_apply(sample)
        ]
        if not candidates:
            raise RuntimeError(f"No eligible sample for {label}: {category}")
        sample = rng.choice(candidates)
        selected.append(sample)
        selected_by_category[category] += 1
        used_pairs.add(sample.pair)
        apply(sample)
        return sample

    # Bring command-line rule coverage to its 100% ceiling.  C1/C2 are used
    # first because they also help the remaining rule-level targets.
    category_order = [(True, True, True), (True, True, False), (True, False, False), (False, False, False)]
    for key in [key for key, state in states.items() if not state.has_commandline]:
        for category in category_order:
            if selected_by_category[category] >= CATEGORY_QUOTAS[category]:
                continue
            try:
                choose(category, lambda sample, key=key: sample.key == key, f"commandline coverage for {key}")
                break
            except RuntimeError:
                continue
        else:
            raise RuntimeError(f"No peer command line available for rule {key}")

    # Reach the executable target first.  This still leaves capacity for the
    # bypass rule targets; reversing the order can strand executable-false
    # rules once their bypass flags have already reached their caps.
    while counts["executable"] < targets["executable"]:
        for category in [(True, False, False), (True, True, False), (True, True, True)]:
            if selected_by_category[category] >= CATEGORY_QUOTAS[category]:
                continue
            try:
                choose(
                    category,
                    lambda sample: not states[sample.key].executable,
                    "rule-level executable coverage",
                )
                break
            except RuntimeError:
                continue
        else:
            raise RuntimeError("Unable to reach the executable rule target")
    while counts["bypass_all"] < targets["bypass_all"]:
        choose(
            (True, True, True),
            lambda sample: not states[sample.key].bypass_all,
            "rule-level bypass-all coverage",
        )
    while counts["bypass_target"] < targets["bypass_target"]:
        try:
            choose(
                (True, True, False),
                lambda sample: not states[sample.key].bypass_target,
                "rule-level bypass-target coverage",
            )
        except RuntimeError:
            choose(
                (True, True, True),
                lambda sample: not states[sample.key].bypass_target,
                "rule-level bypass-target coverage",
            )

    # Fill category quotas using rows that leave the completed rule targets unchanged.
    for category, quota in CATEGORY_QUOTAS.items():
        while selected_by_category[category] < quota:
            choose(
                category,
                lambda sample: (
                    (not sample.executable or states[sample.key].executable)
                    and (not sample.bypass_target or states[sample.key].bypass_target)
                    and (not sample.bypass_all or states[sample.key].bypass_all)
                ),
                "category quota",
            )

    if len(selected) != sum(CATEGORY_QUOTAS.values()):
        raise RuntimeError("The augmentation selection does not match the planned size")
    if counts != targets:
        raise RuntimeError(f"Rule-level targets not met: {counts} != {targets}")
    return selected


def write_workbook(source: Path, destination: Path, samples: list[CommandlineSample]) -> None:
    """Copy Qwen32B and append all selected peer samples with provenance."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    workbook = load_workbook(destination)
    sheets = {sheet.title: sheet for sheet in workbook.worksheets}
    source_headers: dict[str, list[str]] = {}

    for sheet in workbook.worksheets:
        headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if headers is None:
            continue
        source_headers[sheet.title] = [text(value) for value in headers]
        if REQUIRED_HEADERS.issubset(header_map(headers)) and "Supplement_source" not in source_headers[sheet.title]:
            column = sheet.max_column + 1
            header_cell = sheet.cell(1, column, "Supplement_source")
            if column > 1:
                header_cell._style = copy(sheet.cell(1, column - 1)._style)
                header_cell.font = copy(sheet.cell(1, column - 1).font)
            source_headers[sheet.title].append("Supplement_source")

    for sample in samples:
        sheet = sheets[sample.tactic]
        headers = source_headers[sample.tactic]
        values = [sample.values.get(header) if header != "Supplement_source" else sample.source_model for header in headers]
        sheet.append(values)

    workbook.save(destination)


def write_audit(destination: Path, samples: list[CommandlineSample]) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "seed",
        "source_model",
        "tactic",
        "rule_id",
        "rule_name",
        "commandline_evasion",
        "executable",
        "bypass_target_rule",
        "bypass_all_rule",
    ]
    with destination.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for sample in samples:
            writer.writerow(
                {
                    "seed": SEED,
                    "source_model": sample.source_model,
                    "tactic": sample.tactic,
                    "rule_id": sample.rule_id,
                    "rule_name": sample.rule_name,
                    "commandline_evasion": sample.commandline,
                    "executable": sample.executable,
                    "bypass_target_rule": sample.bypass_target,
                    "bypass_all_rule": sample.bypass_all,
                }
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-root",
        type=Path,
        default=PROJECT_ROOT / "results" / "augmented" / "qwen32b",
        help="Directory for the augmented Qwen32B artifacts.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    all_sources = [QWEN32_SOURCE, *PEER_SOURCES.values()]
    missing = [str(path) for path in all_sources if not path.is_file()]
    if missing:
        raise FileNotFoundError("Missing source workbooks:\n" + "\n".join(missing))

    qwen32_records = read_commandlines(QWEN32_SOURCE, "qwen32b")
    peer_records = [record for model, path in PEER_SOURCES.items() for record in read_commandlines(path, model)]
    evidence = collect_evidence(QWEN32_SOURCE)
    states = {
        key: RuleState(
            has_commandline=item.has_commandline,
            executable=False,
            bypass_target=item.has_bypass_target,
            bypass_all=item.has_bypass_all,
        )
        for key, item in evidence.items()
    }
    for record in qwen32_records:
        states[record.key].executable |= record.executable

    before = state_counts(states)
    samples = build_selection(qwen32_records, peer_records, states)
    destination = args.output_root / "qwen32b_res_augmented.xlsx"
    write_workbook(QWEN32_SOURCE, destination, samples)
    write_audit(args.output_root / "added_commandlines_audit.csv", samples)
    summary = {
        "seed": SEED,
        "source_workbook": str(QWEN32_SOURCE),
        "peer_workbooks": {model: str(path) for model, path in PEER_SOURCES.items()},
        "added_commandlines": len(samples),
        "added_by_source": {model: sum(sample.source_model == model for sample in samples) for model in PEER_SOURCES},
        "added_by_category": {
            f"executable={category[0]}, bypass_target={category[1]}, bypass_all={category[2]}": sum(
                sample.category == category for sample in samples
            )
            for category in CATEGORY_QUOTAS
        },
        "rule_state_before": before,
        "rule_state_after": state_counts(states),
    }
    (args.output_root / "augmentation_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Added {len(samples)} command lines to {destination}")
    print(f"Rule-level state: {before} -> {state_counts(states)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
