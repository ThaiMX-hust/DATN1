"""Create filtered result workbooks using evidence from all three models.

A rule is removed only from a model workbook where it has no non-empty
``Commandline_evasion`` value.  It remains eligible only when neither
``Bypass target rule`` nor ``Bypass all rule`` is true for that same rule in
any of the three model workbooks.  Rules are matched by Rule_id, falling back
to Rule_name when the ID is empty.

The source workbooks are never changed.  Filtered copies and a CSV audit are
written beneath the chosen output root.
"""

from __future__ import annotations

import argparse
import csv
import shutil
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TRUE_VALUES = {"true", "1", "1.0", "yes", "y"}
REQUIRED_HEADERS = {
    "Rule_name",
    "Rule_id",
    "Commandline_evasion",
    "Bypass target rule",
    "Bypass all rule",
}
MODEL_WORKBOOKS = {
    "qwen32b": PROJECT_ROOT / "results" / "qwen32b" / "qwen32b_res.xlsx",
    "qwen14b": PROJECT_ROOT / "results" / "qwen14b" / "qwen14b_res_by_tactic.xlsx",
    "gemini": PROJECT_ROOT / "results" / "gemini" / "gemini_res.xlsx",
}


def text(value: Any) -> str:
    """Return a normalized value suitable for matching workbook cells."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace("\r\n", "\n")


def is_true(value: Any) -> bool:
    return value is True or text(value).casefold() in TRUE_VALUES


def rule_identity(rule_id: Any, rule_name: Any) -> str:
    normalized_id = text(rule_id)
    if normalized_id:
        return f"id:{normalized_id}"
    normalized_name = text(rule_name)
    return f"name:{normalized_name}" if normalized_name else ""


def header_map(headers: tuple[Any, ...]) -> dict[str, int]:
    return {text(value): index for index, value in enumerate(headers) if text(value)}


@dataclass
class RuleEvidence:
    rule_id: str = ""
    rule_name: str = ""
    has_commandline: bool = False
    has_bypass_target: bool = False
    has_bypass_all: bool = False
    tactics: set[str] = field(default_factory=set)
    source_rows: int = 0


def collect_evidence(workbook_path: Path) -> dict[str, RuleEvidence]:
    """Collect rule-level evidence from every valid tactic sheet."""
    evidence: dict[str, RuleEvidence] = defaultdict(RuleEvidence)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            continue
        columns = header_map(headers)
        if not REQUIRED_HEADERS.issubset(columns):
            continue

        for row in rows:
            if not any(value not in (None, "") for value in row):
                continue
            key = rule_identity(row[columns["Rule_id"]], row[columns["Rule_name"]])
            if not key:
                continue
            item = evidence[key]
            item.rule_id = text(row[columns["Rule_id"]])
            item.rule_name = text(row[columns["Rule_name"]])
            item.has_commandline |= bool(text(row[columns["Commandline_evasion"]]))
            item.has_bypass_target |= is_true(row[columns["Bypass target rule"]])
            item.has_bypass_all |= is_true(row[columns["Bypass all rule"]])
            item.tactics.add(worksheet.title)
            item.source_rows += 1

    return dict(evidence)


def deletion_sets(
    evidence_by_model: dict[str, dict[str, RuleEvidence]],
) -> dict[str, set[str]]:
    """Return eligible rules, scoped to the workbook where they lack a CLI."""
    all_rule_keys = set().union(*(evidence.keys() for evidence in evidence_by_model.values()))
    globally_no_bypass = {
        key
        for key in all_rule_keys
        if not any(
            item.has_bypass_target or item.has_bypass_all
            for evidence in evidence_by_model.values()
            if (item := evidence.get(key)) is not None
        )
    }
    return {
        model: {
            key
            for key, item in evidence.items()
            if not item.has_commandline and key in globally_no_bypass
        }
        for model, evidence in evidence_by_model.items()
    }


def write_filtered_workbook(source: Path, destination: Path, keys_to_delete: set[str]) -> dict[str, int]:
    """Copy a workbook and remove all rows belonging to eligible rules."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    workbook = load_workbook(destination)
    deleted_rows_by_sheet: dict[str, int] = {}

    for worksheet in workbook.worksheets:
        headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if headers is None:
            continue
        columns = header_map(headers)
        if not {"Rule_name", "Rule_id"}.issubset(columns):
            continue

        rows_to_delete: list[int] = []
        for row_index in range(2, worksheet.max_row + 1):
            key = rule_identity(
                worksheet.cell(row_index, columns["Rule_id"] + 1).value,
                worksheet.cell(row_index, columns["Rule_name"] + 1).value,
            )
            if key in keys_to_delete:
                rows_to_delete.append(row_index)
        for row_index in reversed(rows_to_delete):
            worksheet.delete_rows(row_index, 1)
        if rows_to_delete:
            deleted_rows_by_sheet[worksheet.title] = len(rows_to_delete)

    workbook.save(destination)
    return deleted_rows_by_sheet


def write_audit(
    destination: Path,
    evidence_by_model: dict[str, dict[str, RuleEvidence]],
    keys_by_model: dict[str, set[str]],
) -> None:
    """Write a traceable list of rules removed from each workbook."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "model_removed_from",
        "rule_id",
        "rule_name",
        "tactics_removed_from",
        "source_rows_removed",
        "qwen32b_has_commandline",
        "qwen14b_has_commandline",
        "gemini_has_commandline",
        "qwen32b_bypass_target_or_all_true",
        "qwen14b_bypass_target_or_all_true",
        "gemini_bypass_target_or_all_true",
    ]
    with destination.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for model in MODEL_WORKBOOKS:
            for key in sorted(keys_by_model[model]):
                item = evidence_by_model[model][key]
                row = {
                    "model_removed_from": model,
                    "rule_id": item.rule_id,
                    "rule_name": item.rule_name,
                    "tactics_removed_from": "; ".join(sorted(item.tactics)),
                    "source_rows_removed": item.source_rows,
                }
                for other_model in MODEL_WORKBOOKS:
                    other = evidence_by_model[other_model].get(key)
                    row[f"{other_model}_has_commandline"] = "" if other is None else other.has_commandline
                    row[f"{other_model}_bypass_target_or_all_true"] = (
                        "" if other is None else other.has_bypass_target or other.has_bypass_all
                    )
                writer.writerow(row)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-root",
        type=Path,
        default=PROJECT_ROOT / "results" / "filtered",
        help="Directory for filtered workbook copies (default: results/filtered).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    sources = {model: path.resolve() for model, path in MODEL_WORKBOOKS.items()}
    missing = [str(path) for path in sources.values() if not path.is_file()]
    if missing:
        raise FileNotFoundError("Missing source workbooks:\n" + "\n".join(missing))

    evidence_by_model = {model: collect_evidence(path) for model, path in sources.items()}
    keys_by_model = deletion_sets(evidence_by_model)
    audit_path = args.output_root / "removed_rules_audit.csv"
    write_audit(audit_path, evidence_by_model, keys_by_model)

    for model, source in sources.items():
        destination = args.output_root / model / f"{source.stem}_filtered.xlsx"
        deleted_rows = write_filtered_workbook(source, destination, keys_by_model[model])
        print(
            f"{model}: removed {len(keys_by_model[model])} rules / "
            f"{sum(deleted_rows.values())} rows -> {destination}"
        )
    print(f"Audit written to: {audit_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
