"""Import result CSV rows into a tactic-based final_result.xlsx workbook."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from summarize_rules_by_tactic import iter_rule_files, read_rule_metadata


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_CSV = PROJECT_ROOT / "results" / "gemini" / "gemini_fail_rerun - result.csv"
DEFAULT_WORKBOOK = PROJECT_ROOT / "results" / "gemini" / "final_result.xlsx"
DEFAULT_RULES_DIR = PROJECT_ROOT / "rules"
DEFAULT_REPORT_DIR = PROJECT_ROOT / "output" / "final_result_imports"

SUMMARY_SHEET_NAMES = {"fill_summary", "summary"}
BOOL_COLUMNS = {"Excutable", "Bypass target rule", "Bypass all rule"}
REQUIRED_INPUT_COLUMNS = {"Rule_name"}
TRUE_VALUES = {"true", "1", "1.0", "yes", "y"}
FALSE_VALUES = {"false", "0", "0.0", "no", "n"}


@dataclass(frozen=True)
class WorkbookSheet:
    """A workbook sheet that corresponds to one ATT&CK tactic."""

    title: str
    tactic: str


@dataclass(frozen=True)
class ImportAssignment:
    """One CSV row assigned to one tactic sheet."""

    input_row_number: int
    tactic: str
    sheet_name: str
    mapping_source: str
    csv_row: dict[str, str]


def normalize_token(value: Any) -> str:
    """Normalize a label for matching sheet names and tactic tags."""
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def normalize_technique(value: Any) -> str:
    """Normalize a technique value such as attack.t1059.001 or T1059.001."""
    text = str(value or "").strip().lower()
    if text.startswith("attack."):
        text = text.removeprefix("attack.")
    return text


def parse_scalar(value: str | None) -> str | int | float | None:
    """Parse a CLI scalar into a small Excel-friendly value."""
    if value is None:
        return None
    text = value.strip()
    if not text:
        return None
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)
    return text


def cell_text(value: Any) -> str:
    """Return a normalized text form used for duplicate checks."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_csv_rows(input_csv: Path) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    """Read the input CSV and keep original worksheet row numbers."""
    with input_csv.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        headers = reader.fieldnames or []
        missing = sorted(REQUIRED_INPUT_COLUMNS - set(headers))
        if missing:
            raise ValueError(f"input CSV is missing required column(s): {', '.join(missing)}")
        rows = [(index, dict(row)) for index, row in enumerate(reader, start=2)]
    return headers, rows


def rule_mapping(rules_dir: Path) -> tuple[dict[str, set[str]], dict[str, set[str]]]:
    """Return rule->tactics and technique->tactics maps from Sigma rule tags."""
    by_rule: dict[str, set[str]] = defaultdict(set)
    by_technique: dict[str, set[str]] = defaultdict(set)

    for rule_path in iter_rule_files(rules_dir):
        metadata = read_rule_metadata(rule_path)
        for tactic in metadata.tactics:
            by_rule[metadata.rule_name].add(tactic)
            for technique in metadata.techniques:
                by_technique[normalize_technique(technique)].add(tactic)

    return by_rule, by_technique


def workbook_tactic_sheets(workbook: Any) -> dict[str, WorkbookSheet]:
    """Map tactic slugs to workbook sheets with matching tactic names."""
    sheets: dict[str, WorkbookSheet] = {}
    for worksheet in workbook.worksheets:
        tactic = normalize_token(worksheet.title)
        if tactic in SUMMARY_SHEET_NAMES:
            continue
        sheets[tactic] = WorkbookSheet(title=worksheet.title, tactic=tactic)
    return sheets


def header_map(worksheet: Any) -> dict[str, int]:
    """Return header text to one-based column index for a worksheet."""
    headers: dict[str, int] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(1, column).value
        if value not in (None, ""):
            headers[str(value).strip()] = column
    return headers


def existing_rule_sheet_map(workbook: Any, tactic_sheets: dict[str, WorkbookSheet]) -> dict[str, set[str]]:
    """Return rule names already present in each workbook tactic sheet."""
    rule_to_tactics: dict[str, set[str]] = defaultdict(set)
    sheet_titles = {sheet.title: tactic for tactic, sheet in tactic_sheets.items()}

    for worksheet in workbook.worksheets:
        tactic = sheet_titles.get(worksheet.title)
        if not tactic:
            continue
        headers = header_map(worksheet)
        rule_col = headers.get("Rule_name")
        if rule_col is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            rule_name = worksheet.cell(row, rule_col).value
            if rule_name not in (None, ""):
                rule_to_tactics[str(rule_name).strip()].add(tactic)

    return rule_to_tactics


def resolve_tactics(
    csv_row: dict[str, str],
    by_rule: dict[str, set[str]],
    by_existing_workbook: dict[str, set[str]],
    by_technique: dict[str, set[str]],
    fallback_by_technique: bool,
) -> tuple[list[str], str]:
    """Resolve the tactics for one CSV row."""
    rule_name = str(csv_row.get("Rule_name") or "").strip()
    if rule_name in by_rule:
        return sorted(by_rule[rule_name]), "rule_tags"
    if rule_name in by_existing_workbook:
        return sorted(by_existing_workbook[rule_name]), "existing_workbook"
    if fallback_by_technique:
        technique = normalize_technique(csv_row.get("Technical"))
        if technique in by_technique:
            return sorted(by_technique[technique]), "technique_fallback"
    return [], "unmapped"


def coerce_cell_value(header: str, value: Any) -> Any:
    """Coerce CSV text into a value that matches the workbook convention."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        if header in BOOL_COLUMNS:
            lowered = text.lower()
            if lowered in TRUE_VALUES:
                return True
            if lowered in FALSE_VALUES:
                return False
        return value
    return value


def row_values_for_headers(
    csv_row: dict[str, str],
    headers: dict[str, int],
    llm_value: Any,
) -> dict[int, Any]:
    """Return cell values keyed by one-based workbook column index."""
    values: dict[int, Any] = {}
    for header, column in headers.items():
        if header == "LLM" and llm_value is not None:
            values[column] = llm_value
        elif header in csv_row:
            values[column] = coerce_cell_value(header, csv_row.get(header))
        else:
            values[column] = None
    return values


def last_value_row(worksheet: Any) -> int:
    """Return the last row that has at least one real value."""
    for row in range(worksheet.max_row, 0, -1):
        for column in range(1, worksheet.max_column + 1):
            if worksheet.cell(row, column).value not in (None, ""):
                return row
    return 1


def copy_row_style(worksheet: Any, source_row: int, target_row: int) -> None:
    """Copy row and cell styles from one row to another."""
    if source_row <= 0 or source_row == target_row:
        return

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel

    for column in range(1, worksheet.max_column + 1):
        source_cell = worksheet.cell(source_row, column)
        target_cell = worksheet.cell(target_row, column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def extend_tables(worksheet: Any, new_last_row: int, max_column: int) -> list[dict[str, str]]:
    """Extend Excel tables so appended rows are included."""
    updates: list[dict[str, str]] = []
    for table_name in list(worksheet.tables):
        table = worksheet.tables[table_name]
        min_col, min_row, table_max_col, table_max_row = range_boundaries(table.ref)
        if min_row != 1 or new_last_row <= table_max_row:
            continue

        old_ref = table.ref
        end_col = max(table_max_col, max_column)
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(end_col)}{new_last_row}"
        if table.autoFilter is not None:
            table.autoFilter.ref = table.ref
        updates.append({"sheet": worksheet.title, "table": table_name, "old_ref": old_ref, "new_ref": table.ref})
    return updates


def existing_keys_for_sheet(worksheet: Any, headers: dict[str, int]) -> set[tuple[str, str, str, str, str]]:
    """Build duplicate-detection keys for one sheet."""
    keys: set[tuple[str, str, str, str, str]] = set()
    rule_col = headers.get("Rule_name")
    rule_id_col = headers.get("Rule_id")
    command_col = headers.get("Commandline_evasion")
    llm_col = headers.get("LLM")
    if rule_col is None:
        return keys

    for row in range(2, worksheet.max_row + 1):
        rule_name = worksheet.cell(row, rule_col).value
        if rule_name in (None, ""):
            continue
        key = (
            worksheet.title,
            cell_text(rule_name),
            cell_text(worksheet.cell(row, rule_id_col).value if rule_id_col else ""),
            cell_text(worksheet.cell(row, command_col).value if command_col else ""),
            cell_text(worksheet.cell(row, llm_col).value if llm_col else ""),
        )
        keys.add(key)
    return keys


def key_for_values(sheet_name: str, values: dict[int, Any], headers: dict[str, int]) -> tuple[str, str, str, str, str]:
    """Return the duplicate key for a pending appended row."""
    return (
        sheet_name,
        cell_text(values.get(headers.get("Rule_name", -1))),
        cell_text(values.get(headers.get("Rule_id", -1))),
        cell_text(values.get(headers.get("Commandline_evasion", -1))),
        cell_text(values.get(headers.get("LLM", -1))),
    )


def safe_report_name(input_csv: Path, dry_run: bool) -> str:
    """Return a filesystem-safe report name."""
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", input_csv.stem).strip("_") or "import"
    suffix = "dry_run" if dry_run else "import"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{stem}_{suffix}_{timestamp}.json"


def build_assignments(
    csv_rows: list[tuple[int, dict[str, str]]],
    tactic_sheets: dict[str, WorkbookSheet],
    by_rule: dict[str, set[str]],
    by_existing_workbook: dict[str, set[str]],
    by_technique: dict[str, set[str]],
    fallback_by_technique: bool,
) -> tuple[list[ImportAssignment], list[dict[str, Any]], Counter[str]]:
    """Assign CSV rows to workbook tactic sheets."""
    assignments: list[ImportAssignment] = []
    skipped: list[dict[str, Any]] = []
    mapping_sources: Counter[str] = Counter()

    for row_number, csv_row in csv_rows:
        tactics, source = resolve_tactics(
            csv_row,
            by_rule,
            by_existing_workbook,
            by_technique,
            fallback_by_technique,
        )
        mapping_sources[source] += 1
        if not tactics:
            skipped.append(
                {
                    "input_row": row_number,
                    "rule_name": csv_row.get("Rule_name"),
                    "technical": csv_row.get("Technical"),
                    "reason": "no tactic mapping found",
                }
            )
            continue

        for tactic in tactics:
            sheet = tactic_sheets.get(tactic)
            if sheet is None:
                skipped.append(
                    {
                        "input_row": row_number,
                        "rule_name": csv_row.get("Rule_name"),
                        "technical": csv_row.get("Technical"),
                        "tactic": tactic,
                        "reason": "workbook has no sheet for tactic",
                    }
                )
                continue
            assignments.append(
                ImportAssignment(
                    input_row_number=row_number,
                    tactic=tactic,
                    sheet_name=sheet.title,
                    mapping_source=source,
                    csv_row=csv_row,
                )
            )

    return assignments, skipped, mapping_sources


def import_assignments(
    workbook: Any,
    assignments: list[ImportAssignment],
    llm_value: Any,
    allow_duplicates: bool,
    dry_run: bool = False,
) -> tuple[Counter[str], int, list[dict[str, Any]], list[dict[str, str]]]:
    """Append assignments to workbook sheets, or plan them in dry-run mode."""
    sheet_headers = {worksheet.title: header_map(worksheet) for worksheet in workbook.worksheets}
    existing_keys: dict[str, set[tuple[str, str, str, str, str]]] = {}
    for worksheet in workbook.worksheets:
        existing_keys[worksheet.title] = existing_keys_for_sheet(worksheet, sheet_headers[worksheet.title])

    appended_by_sheet: Counter[str] = Counter()
    skipped_duplicates: list[dict[str, Any]] = []
    table_updates: list[dict[str, str]] = []
    new_last_rows: dict[str, int] = {}

    for assignment in assignments:
        worksheet = workbook[assignment.sheet_name]
        headers = sheet_headers[worksheet.title]
        values = row_values_for_headers(assignment.csv_row, headers, llm_value)
        duplicate_key = key_for_values(worksheet.title, values, headers)
        if not allow_duplicates and duplicate_key in existing_keys[worksheet.title]:
            skipped_duplicates.append(
                {
                    "input_row": assignment.input_row_number,
                    "sheet": worksheet.title,
                    "rule_name": assignment.csv_row.get("Rule_name"),
                    "reason": "duplicate sheet/rule/rule_id/commandline/llm",
                }
            )
            continue

        if dry_run:
            existing_keys[worksheet.title].add(duplicate_key)
            appended_by_sheet[worksheet.title] += 1
            continue

        source_row = max(last_value_row(worksheet), 1)
        target_row = source_row + 1
        copy_row_style(worksheet, source_row, target_row)
        for column, value in values.items():
            worksheet.cell(target_row, column).value = value

        existing_keys[worksheet.title].add(duplicate_key)
        appended_by_sheet[worksheet.title] += 1
        new_last_rows[worksheet.title] = max(new_last_rows.get(worksheet.title, 0), target_row)

    if not dry_run:
        for sheet_name, new_last_row in new_last_rows.items():
            worksheet = workbook[sheet_name]
            table_updates.extend(extend_tables(worksheet, new_last_row, worksheet.max_column))

    return appended_by_sheet, sum(appended_by_sheet.values()), skipped_duplicates, table_updates


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(
        description="Append result CSV rows into tactic sheets in final_result.xlsx."
    )
    parser.add_argument("--input-csv", type=Path, default=DEFAULT_INPUT_CSV)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--rules-dir", type=Path, default=DEFAULT_RULES_DIR)
    parser.add_argument(
        "--llm-value",
        default=None,
        help="Value to write into the workbook LLM column, for example 3 for gemini_add.",
    )
    parser.add_argument("--report", type=Path, help="Report JSON path. Defaults to output/final_result_imports/.")
    parser.add_argument("--dry-run", action="store_true", help="Only report what would be appended.")
    parser.add_argument("--allow-duplicates", action="store_true", help="Append rows even when exact keys already exist.")
    parser.add_argument(
        "--fallback-by-technique",
        action="store_true",
        help="If a rule is missing from rules and workbook, map by Technical/technique tags from other rules.",
    )
    parser.add_argument("--no-backup", action="store_true", help="Do not create a .backup-*.xlsx before saving.")
    return parser.parse_args()


def main() -> int:
    """Run the import."""
    args = parse_args()
    input_csv = args.input_csv.resolve()
    workbook_path = args.workbook.resolve()
    rules_dir = args.rules_dir.resolve()
    llm_value = parse_scalar(args.llm_value)

    if not input_csv.exists():
        raise FileNotFoundError(f"input CSV not found: {input_csv}")
    if not workbook_path.exists():
        raise FileNotFoundError(f"workbook not found: {workbook_path}")
    if not rules_dir.exists():
        raise FileNotFoundError(f"rules directory not found: {rules_dir}")

    csv_headers, csv_rows = read_csv_rows(input_csv)
    by_rule, by_technique = rule_mapping(rules_dir)
    workbook = load_workbook(workbook_path)
    tactic_sheets = workbook_tactic_sheets(workbook)
    by_existing_workbook = existing_rule_sheet_map(workbook, tactic_sheets)

    assignments, skipped_unmapped, mapping_sources = build_assignments(
        csv_rows,
        tactic_sheets,
        by_rule,
        by_existing_workbook,
        by_technique,
        args.fallback_by_technique,
    )

    assignment_counts = Counter(assignment.sheet_name for assignment in assignments)
    report_path = (
        args.report.resolve()
        if args.report
        else (DEFAULT_REPORT_DIR / safe_report_name(input_csv, args.dry_run)).resolve()
    )

    backup_path: Path | None = None
    appended_by_sheet: Counter[str] = Counter()
    appended_total = 0
    skipped_duplicates: list[dict[str, Any]] = []
    table_updates: list[dict[str, str]] = []

    if not args.dry_run and not args.no_backup:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = workbook_path.with_name(f"{workbook_path.stem}.backup-{timestamp}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup_path)

    appended_by_sheet, appended_total, skipped_duplicates, table_updates = import_assignments(
        workbook,
        assignments,
        llm_value,
        args.allow_duplicates,
        dry_run=args.dry_run,
    )

    if not args.dry_run:
        workbook.save(workbook_path)

    report = {
        "dry_run": args.dry_run,
        "input_csv": str(input_csv),
        "workbook": str(workbook_path),
        "backup": str(backup_path) if backup_path else None,
        "rules_dir": str(rules_dir),
        "csv_headers": csv_headers,
        "llm_value": llm_value,
        "input_rows": len(csv_rows),
        "assignment_rows": len(assignments),
        "assignment_rows_by_sheet": dict(sorted(assignment_counts.items())),
        "mapping_sources_by_input_row": dict(sorted(mapping_sources.items())),
        "appended_rows": appended_total,
        "appended_rows_by_sheet": dict(sorted(appended_by_sheet.items())),
        "skipped_unmapped_rows": len(skipped_unmapped),
        "skipped_unmapped_examples": skipped_unmapped[:50],
        "skipped_duplicate_rows": len(skipped_duplicates),
        "skipped_duplicate_examples": skipped_duplicates[:50],
        "table_updates": table_updates,
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    action = "Would append" if args.dry_run else "Appended"
    print(f"Read {len(csv_rows)} CSV rows from {input_csv}")
    print(f"{action} {appended_total} rows across {len(appended_by_sheet)} tactic sheets")
    print(f"Skipped unmapped rows: {len(skipped_unmapped)}")
    print(f"Skipped duplicate rows: {len(skipped_duplicates)}")
    if backup_path:
        print(f"Backup: {backup_path}")
    print(f"Report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
