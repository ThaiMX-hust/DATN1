"""Merge final result workbooks and summarize the combined unique results."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOKS = [
    PROJECT_ROOT / "results" / "results_final" / "gemini" / "gemini_res_final.xlsx",
    PROJECT_ROOT / "results" / "results_final" / "qwen14b" / "qwen14b_res_final.xlsx",
    PROJECT_ROOT / "results" / "results_final" / "qwen32b" / "qwen32b_res_final.xlsx",
]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "final_result_stats" / "combined"
DEFAULT_MERGED_WORKBOOK = PROJECT_ROOT / "results" / "results_final" / "combined" / "combined_res_final.xlsx"

SUMMARY_SHEETS = {"fill_summary", "summary", "thong_ke", "thong ke"}
TRUE_VALUES = {"true", "1", "1.0", "yes", "y"}
FALSE_VALUES = {"false", "0", "0.0", "no", "n"}

SUMMARY_FIELDNAMES = [
    "scope",
    "tactic",
    "source_rows",
    "commandline_total",
    "commandline_executable_success",
    "commandline_behavior",
    "commandline_bypass_target_rule",
    "commandline_bypass_all_rule",
    "rule_total",
    "rule_with_commandline",
    "rule_with_executable_success",
    "rule_with_behavior",
    "rule_with_bypass_target_rule",
    "rule_with_bypass_all_rule",
]

COMMANDLINE_FIELDNAMES = [
    "scope",
    "tactic",
    "commandline_total",
    "commandline_executable_success",
    "commandline_behavior",
    "commandline_bypass_target_rule",
    "commandline_bypass_all_rule",
]

RULE_FIELDNAMES = [
    "scope",
    "tactic",
    "rule_total",
    "rule_with_commandline",
    "rule_with_executable_success",
    "rule_with_behavior",
    "rule_with_bypass_target_rule",
    "rule_with_bypass_all_rule",
]

MERGED_WORKBOOK_HEADERS = [
    "Technical",
    "Rule_name",
    "Title",
    "Rule_id",
    "Command_match_rule",
    "Commandline_evasion",
    "Excutable",
    "Bypass target rule",
    "Bypass all rule",
    "behavior",
    "Trigger rule",
    "Source_models",
    "Source_rows",
]


@dataclass
class ResultRecord:
    """One command-line row from one source workbook."""

    source_model: str
    source_workbook: str
    sheet: str
    rule_name: str
    rule_id: str
    commandline: str
    technical: str = ""
    title: str = ""
    command_match_rule: str = ""
    trigger_rule: str = ""
    executable: bool = False
    behavior: bool = False
    bypass_target: bool = False
    bypass_all: bool = False
    source_rows: int = 1
    source_models: set[str] = field(default_factory=set)
    source_workbooks: set[str] = field(default_factory=set)

    def __post_init__(self) -> None:
        self.source_models.add(self.source_model)
        self.source_workbooks.add(self.source_workbook)

    @property
    def tactic_key(self) -> tuple[str, str, str]:
        """Return a duplicate key inside one tactic."""
        return (self.sheet, self.rule_name, self.commandline)

    @property
    def overall_key(self) -> tuple[str, str]:
        """Return the cross-model duplicate key for overall counts."""
        return (self.rule_name, self.commandline)

    @property
    def rule_key(self) -> str:
        """Return the rule identity used for rule-level statistics."""
        return self.rule_name


def cell_text(value: Any) -> str:
    """Return a stable text value for keys and CSV output."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace("\r\n", "\n")


def normalize_sheet_name(value: Any) -> str:
    """Normalize a sheet name for skip checks."""
    return cell_text(value).lower().replace(" ", "_")


def normalize_bool(value: Any) -> bool:
    """Normalize workbook boolean-ish values."""
    if isinstance(value, bool):
        return value
    text = cell_text(value).lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return False


def header_map(headers: tuple[Any, ...]) -> dict[str, int]:
    """Return header text to zero-based tuple index."""
    return {cell_text(value): index for index, value in enumerate(headers) if cell_text(value)}


def first_value(row: tuple[Any, ...], headers: dict[str, int], names: list[str]) -> Any:
    """Return the first non-empty row value for a list of possible headers."""
    for name in names:
        index = headers.get(name)
        if index is None or index >= len(row):
            continue
        value = row[index]
        if value not in (None, ""):
            return value
    return None


def source_model_from_path(workbook_path: Path) -> str:
    """Return the model name for a workbook path."""
    parent = workbook_path.parent.name
    if parent and parent != "results_final":
        return parent
    return workbook_path.stem.replace("_res_final", "")


def merge_text(existing: str, candidate: str) -> str:
    """Keep the first non-empty text value."""
    return existing or candidate


def merge_records(existing: ResultRecord, incoming: ResultRecord) -> ResultRecord:
    """Merge duplicate command-line records with OR for boolean results."""
    existing.technical = merge_text(existing.technical, incoming.technical)
    existing.title = merge_text(existing.title, incoming.title)
    existing.rule_id = merge_text(existing.rule_id, incoming.rule_id)
    existing.command_match_rule = merge_text(existing.command_match_rule, incoming.command_match_rule)
    existing.trigger_rule = merge_text(existing.trigger_rule, incoming.trigger_rule)
    existing.executable = existing.executable or incoming.executable
    existing.bypass_target = existing.bypass_target or incoming.bypass_target
    existing.bypass_all = existing.bypass_all or incoming.bypass_all
    existing.behavior = existing.behavior or incoming.behavior
    existing.source_rows += incoming.source_rows
    existing.source_models.update(incoming.source_models)
    existing.source_workbooks.update(incoming.source_workbooks)
    return existing


def read_workbook_records(
    workbook_path: Path,
) -> tuple[list[ResultRecord], dict[str, set[str]], Counter[str], list[dict[str, str]]]:
    """Read tactic sheets from one source workbook."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    source_model = source_model_from_path(workbook_path)
    records: list[ResultRecord] = []
    rules_by_sheet: dict[str, set[str]] = {}
    source_rows_by_sheet: Counter[str] = Counter()
    skipped_sheets: list[dict[str, str]] = []

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        if normalize_sheet_name(sheet_name) in SUMMARY_SHEETS:
            continue

        row_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = header_map(next(row_iter))
        except StopIteration:
            skipped_sheets.append({"workbook": str(workbook_path), "sheet": sheet_name, "reason": "empty sheet"})
            continue

        required = {"Rule_name", "Excutable", "Bypass target rule", "Bypass all rule"}
        missing = sorted(required - set(headers))
        if "Commandline_evasion" not in headers and "Command_match_rule" not in headers:
            missing.append("Commandline_evasion or Command_match_rule")
        if missing:
            skipped_sheets.append(
                {
                    "workbook": str(workbook_path),
                    "sheet": sheet_name,
                    "reason": f"missing header(s): {', '.join(missing)}",
                }
            )
            continue

        sheet_rules: set[str] = set()
        for row in row_iter:
            if not any(value not in (None, "") for value in row):
                continue

            source_rows_by_sheet[sheet_name] += 1
            rule_name = cell_text(first_value(row, headers, ["Rule_name"]))
            if not rule_name:
                continue
            sheet_rules.add(rule_name)

            commandline = cell_text(first_value(row, headers, ["Commandline_evasion", "Command_match_rule"]))
            if not commandline:
                continue

            bypass_all = normalize_bool(first_value(row, headers, ["Bypass all rule"]))
            behavior = normalize_bool(first_value(row, headers, ["behavior", "Behavior"])) and bypass_all
            records.append(
                ResultRecord(
                    source_model=source_model,
                    source_workbook=str(workbook_path),
                    sheet=sheet_name,
                    rule_name=rule_name,
                    rule_id=cell_text(first_value(row, headers, ["Rule_id"])),
                    commandline=commandline,
                    technical=cell_text(first_value(row, headers, ["Technical"])),
                    title=cell_text(first_value(row, headers, ["Title"])),
                    command_match_rule=cell_text(first_value(row, headers, ["Command_match_rule"])),
                    trigger_rule=cell_text(first_value(row, headers, ["Trigger rule"])),
                    executable=normalize_bool(first_value(row, headers, ["Excutable", "Executable"])),
                    behavior=behavior,
                    bypass_target=normalize_bool(first_value(row, headers, ["Bypass target rule"])),
                    bypass_all=bypass_all,
                )
            )

        rules_by_sheet[sheet_name] = sheet_rules

    return records, rules_by_sheet, source_rows_by_sheet, skipped_sheets


def deduplicate_records(records: list[ResultRecord], overall: bool) -> list[ResultRecord]:
    """Deduplicate command-line records across source models."""
    grouped: dict[tuple[str, ...], ResultRecord] = {}
    for record in records:
        key = record.overall_key if overall else record.tactic_key
        existing = grouped.get(key)
        if existing is None:
            grouped[key] = record
            continue
        grouped[key] = merge_records(existing, record)
    return list(grouped.values())


def ratio(numerator: int, denominator: int) -> float:
    """Return a safe decimal ratio."""
    return numerator / denominator if denominator else 0.0


def summarize_scope(
    scope: str,
    tactic: str,
    records: list[ResultRecord],
    rule_total: int,
    source_rows: int,
    overall: bool,
) -> dict[str, Any]:
    """Summarize one tactic sheet or the whole merged dataset."""
    commandlines = deduplicate_records(records, overall=overall)
    rules_with_commandline = {record.rule_key for record in commandlines}
    rules_with_executable = {record.rule_key for record in commandlines if record.executable}
    rules_with_behavior = {record.rule_key for record in commandlines if record.behavior}
    rules_with_bypass_target = {record.rule_key for record in commandlines if record.bypass_target}
    rules_with_bypass_all = {record.rule_key for record in commandlines if record.bypass_all}

    return {
        "scope": scope,
        "tactic": tactic,
        "source_rows": source_rows,
        "commandline_total": len(commandlines),
        "commandline_executable_success": sum(1 for record in commandlines if record.executable),
        "commandline_behavior": sum(1 for record in commandlines if record.behavior),
        "commandline_bypass_target_rule": sum(1 for record in commandlines if record.bypass_target),
        "commandline_bypass_all_rule": sum(1 for record in commandlines if record.bypass_all),
        "rule_total": rule_total,
        "rule_with_commandline": len(rules_with_commandline),
        "rule_with_executable_success": len(rules_with_executable),
        "rule_with_behavior": len(rules_with_behavior),
        "rule_with_bypass_target_rule": len(rules_with_bypass_target),
        "rule_with_bypass_all_rule": len(rules_with_bypass_all),
    }


def build_summary(
    records: list[ResultRecord],
    rules_by_sheet: dict[str, set[str]],
    source_rows_by_sheet: Counter[str],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Build per-tactic and overall merged summary rows."""
    per_sheet_rows: list[dict[str, Any]] = []
    for sheet_name in rules_by_sheet:
        sheet_records = [record for record in records if record.sheet == sheet_name]
        per_sheet_rows.append(
            summarize_scope(
                scope="tactic",
                tactic=sheet_name,
                records=sheet_records,
                rule_total=len(rules_by_sheet[sheet_name]),
                source_rows=source_rows_by_sheet[sheet_name],
                overall=False,
            )
        )

    all_rules = set().union(*rules_by_sheet.values()) if rules_by_sheet else set()
    overall_row = summarize_scope(
        scope="overall",
        tactic="ALL",
        records=records,
        rule_total=len(all_rules),
        source_rows=sum(source_rows_by_sheet.values()),
        overall=True,
    )
    return per_sheet_rows, overall_row


def commandline_rows(summary_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Extract commandline-level table rows."""
    return [{field: row[field] for field in COMMANDLINE_FIELDNAMES} for row in summary_rows]


def rule_rows(summary_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Extract rule-level table rows."""
    return [{field: row[field] for field in RULE_FIELDNAMES} for row in summary_rows]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    """Write rows to CSV."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def append_table(workbook: Workbook, sheet_name: str, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    """Append a table to an output workbook sheet."""
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(fieldnames)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append([row.get(field) for field in fieldnames])
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        width = max(len(cell_text(cell.value)) for cell in column_cells) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 42)


def write_excel_summary(
    path: Path,
    per_tactic_rows: list[dict[str, Any]],
    overall_row: dict[str, Any],
    skipped_sheets: list[dict[str, str]],
    sources: list[dict[str, str]],
) -> None:
    """Write an XLSX report with separate summary tables."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    all_summary_rows = [overall_row, *per_tactic_rows]

    append_table(workbook, "Overall", SUMMARY_FIELDNAMES, [overall_row])
    append_table(workbook, "By_Tactic", SUMMARY_FIELDNAMES, per_tactic_rows)
    append_table(workbook, "Commandline", COMMANDLINE_FIELDNAMES, commandline_rows(all_summary_rows))
    append_table(workbook, "Rule", RULE_FIELDNAMES, rule_rows(all_summary_rows))
    append_table(workbook, "Sources", ["model", "workbook"], sources)
    append_table(workbook, "Skipped_Sheets", ["workbook", "sheet", "reason"], skipped_sheets)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_merged_workbook(path: Path, records: list[ResultRecord], rules_by_sheet: dict[str, set[str]]) -> None:
    """Write a deduplicated merged workbook for audit."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for tactic in rules_by_sheet:
        worksheet = workbook.create_sheet(tactic)
        worksheet.append(MERGED_WORKBOOK_HEADERS)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        records_for_sheet = sorted(
            deduplicate_records([record for record in records if record.sheet == tactic], overall=False),
            key=lambda record: (record.rule_name, record.commandline),
        )
        for record in records_for_sheet:
            worksheet.append(
                [
                    record.technical,
                    record.rule_name,
                    record.title,
                    record.rule_id,
                    record.command_match_rule,
                    record.commandline,
                    record.executable,
                    record.bypass_target,
                    record.bypass_all,
                    record.behavior,
                    record.trigger_rule,
                    "; ".join(sorted(record.source_models)),
                    record.source_rows,
                ]
            )
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = {
            "A": 13,
            "B": 44,
            "C": 42,
            "D": 38,
            "E": 50,
            "F": 90,
            "G": 13,
            "H": 18,
            "I": 16,
            "J": 13,
            "K": 62,
            "L": 24,
            "M": 12,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(
        description="Merge final result workbooks and summarize unique command lines across models."
    )
    parser.add_argument(
        "--workbook",
        action="append",
        dest="workbooks",
        type=Path,
        help="Workbook to merge. Repeat for multiple files. Defaults to the three final model workbooks.",
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--merged-workbook",
        type=Path,
        default=DEFAULT_MERGED_WORKBOOK,
        help="Deduplicated merged workbook path. Use an empty string to skip writing it.",
    )
    return parser.parse_args()


def main() -> int:
    """Run the merged final result summarizer."""
    args = parse_args()
    workbook_paths = [path.resolve() for path in args.workbooks] if args.workbooks else DEFAULT_WORKBOOKS
    output_dir = args.output_dir.resolve()
    merged_workbook = args.merged_workbook

    records: list[ResultRecord] = []
    rules_by_sheet: dict[str, set[str]] = {}
    source_rows_by_sheet: Counter[str] = Counter()
    skipped_sheets: list[dict[str, str]] = []
    sources: list[dict[str, str]] = []

    for workbook_path in workbook_paths:
        workbook_path = workbook_path.resolve()
        if not workbook_path.exists():
            raise FileNotFoundError(f"workbook not found: {workbook_path}")

        workbook_records, workbook_rules, workbook_source_rows, workbook_skipped = read_workbook_records(workbook_path)
        records.extend(workbook_records)
        for sheet_name, rules in workbook_rules.items():
            rules_by_sheet.setdefault(sheet_name, set()).update(rules)
        source_rows_by_sheet.update(workbook_source_rows)
        skipped_sheets.extend(workbook_skipped)
        sources.append({"model": source_model_from_path(workbook_path), "workbook": str(workbook_path)})

    per_tactic_rows, overall_row = build_summary(records, rules_by_sheet, source_rows_by_sheet)
    summary_rows = [overall_row, *per_tactic_rows]

    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "by_tactic_summary.csv", SUMMARY_FIELDNAMES, per_tactic_rows)
    write_csv(output_dir / "overall_summary.csv", SUMMARY_FIELDNAMES, [overall_row])
    write_csv(output_dir / "commandline_summary.csv", COMMANDLINE_FIELDNAMES, commandline_rows(summary_rows))
    write_csv(output_dir / "rule_summary.csv", RULE_FIELDNAMES, rule_rows(summary_rows))
    write_excel_summary(
        output_dir / "final_result_statistics.xlsx",
        per_tactic_rows,
        overall_row,
        skipped_sheets,
        sources,
    )

    merged_workbook_path = None
    if str(merged_workbook):
        merged_workbook_path = merged_workbook.resolve()
        write_merged_workbook(merged_workbook_path, records, rules_by_sheet)

    payload = {
        "workbooks": [str(path.resolve()) for path in workbook_paths],
        "dedupe": {
            "by_tactic": ["tactic", "Rule_name", "Commandline_evasion"],
            "overall": ["Rule_name", "Commandline_evasion"],
            "booleans": "OR across duplicate rows",
            "behavior": "behavior AND Bypass all rule",
        },
        "merged_workbook": str(merged_workbook_path) if merged_workbook_path else None,
        "overall": overall_row,
        "by_tactic": per_tactic_rows,
        "skipped_sheets": skipped_sheets,
    }
    (output_dir / "final_result_statistics.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Read {len(records)} commandline rows from {len(workbook_paths)} workbooks")
    print("Dedupe overall key: Rule_name + Commandline_evasion")
    print(f"Tactic sheets: {len(per_tactic_rows)}")
    print(f"Overall commandlines: {overall_row['commandline_total']}")
    print(f"Overall behavior commandlines: {overall_row['commandline_behavior']}")
    print(f"Overall rules with commandline: {overall_row['rule_with_commandline']}")
    print(f"Overall rules with behavior: {overall_row['rule_with_behavior']}")
    print(f"Wrote reports to {output_dir}")
    if merged_workbook_path:
        print(f"Wrote merged workbook to {merged_workbook_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
