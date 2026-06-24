"""Summarize rule and command-line results from a tactic-per-sheet workbook.

Each non-empty ``Commandline_evasion`` cell is counted as one command line for
its tactic.  The overall result removes duplicates from rules that occur in
multiple tactic sheets:

* a rule is identified by ``Rule_id`` (or ``Rule_name`` if it has no ID);
* a command line is identified by its rule and its ``Commandline_evasion``.

If duplicate overall command lines have different boolean results, their
results are merged with OR: the command line is considered true when it is
true in at least one occurrence.
"""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REQUIRED_HEADERS = {
    "Rule_name",
    "Rule_id",
    "Commandline_evasion",
    "Excutable",  # This is the header spelling in the source workbook.
    "Bypass target rule",
    "Bypass all rule",
}
TRUE_VALUES = {"true", "1", "1.0", "yes", "y"}

COMMANDLINE_COLUMNS = [
    ("tactic", "Phạm vi / Tactic"),
    ("commandline_total", "Tổng command-line"),
    ("executable_success", "Thực thi thành công"),
    ("executable_rate", "Tỷ lệ thực thi"),
    ("bypass_target_commandline_count", "Bypass target rule"),
    ("bypass_target_per_total", "Bypass target / total"),
    ("bypass_target_per_success", "Bypass target / success"),
    ("bypass_all_commandline_count", "Bypass all rule"),
    ("bypass_all_per_total", "Bypass all / total"),
    ("bypass_all_per_success", "Bypass all / success"),
]

RULE_COLUMNS = [
    ("tactic", "Phạm vi / Tactic"),
    ("rule_total", "Tổng rule"),
    ("rule_with_commandline", "Rule có command-line"),
    ("rule_with_commandline_rate", "Tỷ lệ"),
    ("rule_with_executable_success", "Rule có command-line thực thi thành công"),
    ("rule_with_executable_success_rate", "Tỷ lệ thực thi thành công"),
    ("rule_with_bypass_target", "Rule có bypass target"),
    ("rule_with_bypass_target_rate", "Tỷ lệ bypass target"),
    ("rule_with_bypass_all", "Rule có bypass all"),
    ("rule_with_bypass_all_rate", "Tỷ lệ bypass all"),
]


@dataclass(frozen=True)
class CommandlineRecord:
    """A row with a non-empty command line."""

    tactic: str
    rule_key: str
    commandline: str
    executable: bool
    bypass_target: bool
    bypass_all: bool

    @property
    def commandline_key(self) -> tuple[str, str]:
        return (self.rule_key, self.commandline)

def text(value: Any) -> str:
    """Convert an Excel value to a stable, non-empty key component."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace("\r\n", "\n")


def as_true(value: Any) -> bool:
    """Interpret common Excel boolean representations."""
    return value is True or text(value).casefold() in TRUE_VALUES


def build_header_map(headers: Iterable[Any]) -> dict[str, int]:
    """Map exact trimmed header names to their zero-based indexes."""
    return {text(header): index for index, header in enumerate(headers) if text(header)}


def rule_identity(rule_id: Any, rule_name: Any) -> str:
    """Return the rule identity used for cross-tactic deduplication."""
    normalized_id = text(rule_id)
    if normalized_id:
        return f"id:{normalized_id}"
    normalized_name = text(rule_name)
    return f"name:{normalized_name}" if normalized_name else ""


def read_workbook(
    workbook_path: Path,
) -> tuple[dict[str, set[str]], dict[str, list[CommandlineRecord]], dict[str, int], list[dict[str, str]]]:
    """Read each valid tactic sheet, preserving row-level tactic data."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rules_by_tactic: dict[str, set[str]] = {}
    records_by_tactic: dict[str, list[CommandlineRecord]] = {}
    source_rows_by_tactic: dict[str, int] = {}
    skipped_sheets: list[dict[str, str]] = []

    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = build_header_map(next(rows))
        except StopIteration:
            skipped_sheets.append({"sheet": worksheet.title, "reason": "Sheet rỗng"})
            continue

        missing = sorted(REQUIRED_HEADERS - headers.keys())
        if missing:
            skipped_sheets.append(
                {"sheet": worksheet.title, "reason": f"Thiếu cột: {', '.join(missing)}"}
            )
            continue

        tactic = worksheet.title
        rules: set[str] = set()
        records: list[CommandlineRecord] = []
        source_rows = 0

        for row in rows:
            if not any(value not in (None, "") for value in row):
                continue
            source_rows += 1

            rule_key = rule_identity(row[headers["Rule_id"]], row[headers["Rule_name"]])
            if rule_key:
                rules.add(rule_key)

            commandline = text(row[headers["Commandline_evasion"]])
            if not rule_key or not commandline:
                continue
            records.append(
                CommandlineRecord(
                    tactic=tactic,
                    rule_key=rule_key,
                    commandline=commandline,
                    executable=as_true(row[headers["Excutable"]]),
                    bypass_target=as_true(row[headers["Bypass target rule"]]),
                    bypass_all=as_true(row[headers["Bypass all rule"]]),
                )
            )

        rules_by_tactic[tactic] = rules
        records_by_tactic[tactic] = records
        source_rows_by_tactic[tactic] = source_rows

    return rules_by_tactic, records_by_tactic, source_rows_by_tactic, skipped_sheets


def deduplicate_overall(records: Iterable[CommandlineRecord]) -> list[CommandlineRecord]:
    """Merge duplicate rule--commandline records from different tactic sheets."""
    merged: dict[tuple[str, str], CommandlineRecord] = {}
    for record in records:
        previous = merged.get(record.commandline_key)
        if previous is None:
            merged[record.commandline_key] = record
            continue
        merged[record.commandline_key] = CommandlineRecord(
            tactic="ALL",
            rule_key=record.rule_key,
            commandline=record.commandline,
            executable=previous.executable or record.executable,
            bypass_target=previous.bypass_target or record.bypass_target,
            bypass_all=previous.bypass_all or record.bypass_all,
        )
    return list(merged.values())


def ratio(numerator: int, denominator: int) -> float:
    """Return a safe decimal ratio for Excel percentage formatting."""
    return numerator / denominator if denominator else 0.0


def summarize(
    scope: str,
    tactic: str,
    rules: set[str],
    records: list[CommandlineRecord],
    source_rows: int,
) -> dict[str, Any]:
    """Calculate the commandline and rule tables requested by the report."""
    commandline_total = len(records)
    rule_total = len(rules)
    executable_success = sum(record.executable for record in records)
    bypass_target_commandline_count = sum(record.bypass_target for record in records)
    bypass_all_commandline_count = sum(record.bypass_all for record in records)

    rules_with_commandline = {record.rule_key for record in records}
    rules_with_executable_success = {record.rule_key for record in records if record.executable}
    rules_with_bypass_target = {record.rule_key for record in records if record.bypass_target}
    rules_with_bypass_all = {record.rule_key for record in records if record.bypass_all}

    return {
        "scope": scope,
        "tactic": tactic,
        "source_rows": source_rows,
        "commandline_total": commandline_total,
        "executable_success": executable_success,
        "executable_rate": ratio(executable_success, commandline_total),
        "bypass_target_commandline_count": bypass_target_commandline_count,
        "bypass_target_per_total": ratio(bypass_target_commandline_count, commandline_total),
        "bypass_target_per_success": ratio(bypass_target_commandline_count, executable_success),
        "bypass_all_commandline_count": bypass_all_commandline_count,
        "bypass_all_per_total": ratio(bypass_all_commandline_count, commandline_total),
        "bypass_all_per_success": ratio(bypass_all_commandline_count, executable_success),
        "rule_total": rule_total,
        "rule_with_commandline": len(rules_with_commandline),
        "rule_with_commandline_rate": ratio(len(rules_with_commandline), rule_total),
        "rule_with_executable_success": len(rules_with_executable_success),
        "rule_with_executable_success_rate": ratio(len(rules_with_executable_success), rule_total),
        "rule_with_bypass_target": len(rules_with_bypass_target),
        "rule_with_bypass_target_rate": ratio(len(rules_with_bypass_target), rule_total),
        "rule_with_bypass_all": len(rules_with_bypass_all),
        "rule_with_bypass_all_rate": ratio(len(rules_with_bypass_all), rule_total),
    }


def write_csv(path: Path, columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
    """Write a UTF-8-with-BOM CSV with the report's display headers."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [label for _, label in columns]
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({label: row[key] for key, label in columns})


def append_table(
    worksheet: Any,
    start_row: int,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
) -> int:
    """Write a formatted report table and return the next available row."""
    keys = [key for key, _ in columns]
    for column, (_, label) in enumerate(columns, start=1):
        cell = worksheet.cell(start_row, column, label)
        cell.font = Font(bold=True)
    for summary in rows:
        worksheet.append([summary[key] for key in keys])

    ratio_columns = {
        index + 1
        for index, (key, _) in enumerate(columns)
        if key.endswith("_rate") or key.endswith("_per_total") or key.endswith("_per_success")
    }
    for row in worksheet.iter_rows(min_row=start_row + 1, max_row=start_row + len(rows)):
        for column_index in ratio_columns:
            row[column_index - 1].number_format = "0.00%"
    return start_row + len(rows) + 1


def fit_columns(worksheet: Any, max_width: int = 48) -> None:
    """Set practical column widths for a report sheet."""
    for column in worksheet.columns:
        width = max(len(text(cell.value)) for cell in column) + 2
        worksheet.column_dimensions[column[0].column_letter].width = min(max(width, 14), max_width)


def write_excel(path: Path, overall: dict[str, Any], by_tactic: list[dict[str, Any]], skipped: list[dict[str, str]]) -> None:
    """Write Excel tables matching the requested commandline/rule layout."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    commandline_sheet = workbook.create_sheet("Commandline")
    commandline_sheet["A1"] = "Mức commandline"
    commandline_sheet["A1"].font = Font(bold=True, size=14)
    append_table(commandline_sheet, 3, COMMANDLINE_COLUMNS, [overall])
    commandline_sheet["A6"] = "Theo tactics"
    commandline_sheet["A6"].font = Font(bold=True, size=14)
    append_table(commandline_sheet, 8, COMMANDLINE_COLUMNS, by_tactic)
    commandline_sheet.freeze_panes = "A9"
    fit_columns(commandline_sheet)

    rule_sheet = workbook.create_sheet("Rule")
    rule_sheet["A1"] = "Mức rule"
    rule_sheet["A1"].font = Font(bold=True, size=14)
    append_table(rule_sheet, 3, RULE_COLUMNS, [overall])
    rule_sheet["A6"] = "Theo tactics"
    rule_sheet["A6"].font = Font(bold=True, size=14)
    append_table(rule_sheet, 8, RULE_COLUMNS, by_tactic)
    rule_sheet.freeze_panes = "A9"
    fit_columns(rule_sheet)

    info = workbook.create_sheet("Notes")
    info.append(["Quy tắc tính"])
    info["A1"].font = Font(bold=True)
    notes = [
        "Mỗi dòng có Commandline_evasion không rỗng được tính là một commandline trong tactic.",
        "Overall khử trùng lặp rule theo Rule_id; nếu Rule_id rỗng thì dùng Rule_name.",
        "Overall khử trùng lặp commandline theo (rule, Commandline_evasion).",
        "Bản ghi trùng có kết quả boolean khác nhau sẽ được gộp theo OR.",
        "Tỷ lệ Bypass target / success và Bypass all / success lấy Thực thi thành công làm mẫu số.",
        "Tỷ lệ ở mức rule lấy Tổng rule làm mẫu số.",
    ]
    for note in notes:
        info.append([note])
    info.column_dimensions["A"].width = 120

    skipped_sheet = workbook.create_sheet("Skipped_Sheets")
    skipped_sheet.append(["Sheet", "Lý do"])
    for cell in skipped_sheet[1]:
        cell.font = Font(bold=True)
    for item in skipped:
        skipped_sheet.append([item["sheet"], item["reason"]])
    skipped_sheet.freeze_panes = "A2"
    skipped_sheet.column_dimensions["A"].width = 28
    skipped_sheet.column_dimensions["B"].width = 70

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Thống kê rule và commandline từ workbook có một tactic trên mỗi sheet."
    )
    parser.add_argument("workbook", type=Path, help="Đường dẫn file .xlsx đầu vào")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=PROJECT_ROOT / "output" / "tactic_statistics/filter/gemini3.5flash",
        help="Thư mục ghi các báo cáo (mặc định: output/tactic_statistics/filter/gemini3.5flash)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    output_dir = args.output_dir.resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Không tìm thấy workbook: {workbook_path}")

    rules_by_tactic, records_by_tactic, source_rows_by_tactic, skipped = read_workbook(workbook_path)
    by_tactic = [
        summarize(
            scope="tactic",
            tactic=tactic,
            rules=rules_by_tactic[tactic],
            records=records_by_tactic[tactic],
            source_rows=source_rows_by_tactic[tactic],
        )
        for tactic in rules_by_tactic
    ]

    all_rules = set().union(*rules_by_tactic.values()) if rules_by_tactic else set()
    all_records = deduplicate_overall(
        record for records in records_by_tactic.values() for record in records
    )
    overall = summarize(
        scope="overall",
        tactic="ALL",
        rules=all_rules,
        records=all_records,
        source_rows=sum(source_rows_by_tactic.values()),
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    all_scopes = [overall, *by_tactic]
    write_csv(output_dir / "commandline_summary.csv", COMMANDLINE_COLUMNS, all_scopes)
    write_csv(output_dir / "rule_summary.csv", RULE_COLUMNS, all_scopes)
    write_excel(output_dir / "tactic_statistics.xlsx", overall, by_tactic, skipped)
    (output_dir / "tactic_statistics.json").write_text(
        json.dumps(
            {
                "input_workbook": str(workbook_path),
                "overall": overall,
                "by_tactic": by_tactic,
                "skipped_sheets": skipped,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    # Keep console output ASCII so the script also runs in legacy Windows shells.
    print(f"Processed {len(by_tactic)} tactic sheets.")
    print(f"Overall: {overall['rule_total']} rules, {overall['commandline_total']} commandlines.")
    print(f"Reports written to: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
