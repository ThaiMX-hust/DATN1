"""Relabel exact Rule/CommandLine matches from a review CSV as non-executable.

The input review file is expected to have ``Rule`` and ``CommandLine``
columns.  Matching is exact after trimming and normalizing line endings.
Only the ``Excutable`` cell is set to the Boolean value ``False``; bypass
columns remain unchanged.  A workbook backup and a full audit are written.
"""

from __future__ import annotations

import argparse
import csv
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from filter_rules_without_commandlines import REQUIRED_HEADERS, header_map, is_true, text


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = (
    PROJECT_ROOT
    / "results"
    / "filtered_common"
    / "gemini"
    / "gemini_res_filtered_common_denoised.xlsx"
)
DEFAULT_REVIEW = Path(
    r"C:\Users\thaim\.codex\attachments\5f4615c1-ae77-4e31-9892-ffc39d7b4d13\pasted-text.txt"
)


def read_review_pairs(path: Path) -> tuple[set[tuple[str, str]], list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as file:
        rows = list(csv.DictReader(file))
    required = {"Rule", "CommandLine"}
    if not rows or not required.issubset(rows[0]):
        raise ValueError(f"Review file must include columns: {', '.join(sorted(required))}")
    return {(text(row["Rule"]), text(row["CommandLine"])) for row in rows}, rows


def apply_labels(
    workbook_path: Path, review_pairs: set[tuple[str, str]]
) -> tuple[list[dict[str, object]], set[tuple[str, str]]]:
    """Set the executable label to False for every exact match."""
    workbook = load_workbook(workbook_path)
    audit: list[dict[str, object]] = []
    matched_pairs: set[tuple[str, str]] = set()

    for worksheet in workbook.worksheets:
        headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if headers is None:
            continue
        columns = header_map(headers)
        if not REQUIRED_HEADERS.issubset(columns):
            continue
        for row_number in range(2, worksheet.max_row + 1):
            pair = (
                text(worksheet.cell(row_number, columns["Rule_name"] + 1).value),
                text(worksheet.cell(row_number, columns["Commandline_evasion"] + 1).value),
            )
            if pair not in review_pairs:
                continue
            executable_cell = worksheet.cell(row_number, columns["Excutable"] + 1)
            previous = is_true(executable_cell.value)
            executable_cell.value = False
            matched_pairs.add(pair)
            audit.append(
                {
                    "tactic": worksheet.title,
                    "row_number": row_number,
                    "rule_name": pair[0],
                    "commandline_evasion": pair[1],
                    "executable_before": previous,
                    "executable_after": False,
                    "changed": previous,
                }
            )
    workbook.save(workbook_path)
    return audit, matched_pairs


def write_audit(
    directory: Path,
    audit: list[dict[str, object]],
    review_pairs: set[tuple[str, str]],
    matched_pairs: set[tuple[str, str]],
) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    with (directory / "gemini_relabel_excutable_false_audit.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as file:
        fields = [
            "tactic",
            "row_number",
            "rule_name",
            "commandline_evasion",
            "executable_before",
            "executable_after",
            "changed",
        ]
        writer = csv.DictWriter(file, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(audit)
    with (directory / "gemini_relabel_unmatched_review_pairs.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as file:
        writer = csv.DictWriter(file, fieldnames=["rule_name", "commandline_evasion"], lineterminator="\n")
        writer.writeheader()
        for rule_name, commandline in sorted(review_pairs - matched_pairs):
            writer.writerow({"rule_name": rule_name, "commandline_evasion": commandline})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--review", type=Path, default=DEFAULT_REVIEW)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    review_path = args.review.resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if not review_path.is_file():
        raise FileNotFoundError(f"Review file not found: {review_path}")

    review_pairs, _ = read_review_pairs(review_path)
    backup_path = workbook_path.with_name(f"{workbook_path.stem}_before_relabel{workbook_path.suffix}")
    if not backup_path.exists():
        shutil.copy2(workbook_path, backup_path)
    audit, matched_pairs = apply_labels(workbook_path, review_pairs)
    write_audit(workbook_path.parent, audit, review_pairs, matched_pairs)
    changed = sum(item["changed"] for item in audit)
    print(f"Matched {len(matched_pairs)} of {len(review_pairs)} review pairs across {len(audit)} rows.")
    print(f"Changed {changed} Excutable labels to False.")
    print(f"Backup: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
