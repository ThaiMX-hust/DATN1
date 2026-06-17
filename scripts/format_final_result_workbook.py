"""Format final_result.xlsx tactic sheets for easier reading."""

from __future__ import annotations

import argparse
import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = PROJECT_ROOT / "results" / "gemini" / "final_result.xlsx"

SUMMARY_SHEET_NAMES = {"fill_summary", "summary"}
TACTIC_REQUIRED_HEADERS = {"Rule_name", "Commandline_evasion", "Excutable", "Bypass target rule", "Bypass all rule"}
BOOL_HEADERS = {"Excutable", "Bypass target rule", "Bypass all rule"}
TRUE_VALUES = {"true", "1", "1.0", "yes", "y"}
FALSE_VALUES = {"false", "0", "0.0", "no", "n"}

TACTIC_WIDTHS = {
    "Technical": 13,
    "Rule_name": 44,
    "Title": 42,
    "Rule_id": 38,
    "Command_match_rule": 50,
    "Commandline_evasion": 90,
    "Excutable": 13,
    "Bypass target rule": 18,
    "Bypass all rule": 16,
    "Trigger rule": 62,
    "LLM": 9,
}
SUMMARY_WIDTHS = {
    "Metric": 42,
    "Value": 18,
    "Note": 52,
    "Source / Sheet": 45,
    "Rows": 44,
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="7030A0")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
ALT_ROW_FILL = PatternFill("solid", fgColor="F7FAFC")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
TRUE_FILL = PatternFill("solid", fgColor="E2F0D9")
FALSE_FILL = PatternFill("solid", fgColor="FCE4D6")
BLANK_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def normalize_sheet_name(value: Any) -> str:
    """Normalize sheet names for summary-sheet checks."""
    return str(value or "").strip().lower().replace(" ", "_")


def cell_text(value: Any) -> str:
    """Return display text for a workbook cell value."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def normalize_bool(value: Any) -> str:
    """Return true, false, or blank for boolean-ish values."""
    text = cell_text(value).strip().lower()
    if text in TRUE_VALUES:
        return "true"
    if text in FALSE_VALUES:
        return "false"
    return ""


def header_map(worksheet: Any) -> dict[str, int]:
    """Return header text to one-based column index."""
    return {
        str(worksheet.cell(1, column).value).strip(): column
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, column).value not in (None, "")
    }


def last_value_row(worksheet: Any) -> int:
    """Return the last row that has at least one real value."""
    for row in range(worksheet.max_row, 0, -1):
        if any(worksheet.cell(row, column).value not in (None, "") for column in range(1, worksheet.max_column + 1)):
            return row
    return 1


def is_tactic_sheet(headers: dict[str, int]) -> bool:
    """Return whether a sheet has the final_result tactic schema."""
    return TACTIC_REQUIRED_HEADERS.issubset(headers)


def set_column_widths(worksheet: Any, widths: dict[str, int]) -> None:
    """Apply column widths using header names."""
    headers = header_map(worksheet)
    for header, column in headers.items():
        width = widths.get(header, 16)
        worksheet.column_dimensions[get_column_letter(column)].width = width


def row_height_for_texts(texts: list[tuple[str, int]], minimum: float = 20.0, maximum: float = 72.0) -> float:
    """Estimate a compact wrapped row height."""
    line_count = 1
    for text, width in texts:
        if not text:
            continue
        explicit_lines = text.count("\n") + 1
        wrapped_lines = max(1, math.ceil(len(text) / max(width, 1)))
        line_count = max(line_count, explicit_lines, wrapped_lines)
    return min(max(minimum, 15.0 * line_count), maximum)


def format_header_row(worksheet: Any, fill: PatternFill) -> None:
    """Format the first row as a readable header."""
    worksheet.row_dimensions[1].height = 28
    for cell in worksheet[1]:
        if cell.value in (None, ""):
            continue
        cell.fill = fill
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def format_bool_cell(cell: Any) -> None:
    """Format a boolean-ish cell with a light status fill."""
    value = normalize_bool(cell.value)
    if value == "true":
        cell.fill = TRUE_FILL
        cell.font = Font(name="Calibri", size=10, color="006100")
    elif value == "false":
        cell.fill = FALSE_FILL
        cell.font = Font(name="Calibri", size=10, color="9C0006")
    else:
        cell.fill = BLANK_FILL
        cell.font = Font(name="Calibri", size=10, color="666666")


def extend_tables_to_last_row(worksheet: Any, last_row: int, last_col: int) -> list[dict[str, str]]:
    """Extend existing Excel tables to include the visible formatted range."""
    updates: list[dict[str, str]] = []
    for table_name in list(worksheet.tables):
        table = worksheet.tables[table_name]
        min_col, min_row, table_max_col, table_max_row = range_boundaries(table.ref)
        new_max_col = max(table_max_col, last_col)
        if table_max_row == last_row and table_max_col == new_max_col:
            continue
        old_ref = table.ref
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(new_max_col)}{last_row}"
        if table.autoFilter is not None:
            table.autoFilter.ref = table.ref
        updates.append({"sheet": worksheet.title, "table": table_name, "old_ref": old_ref, "new_ref": table.ref})
    return updates


def format_tactic_sheet(worksheet: Any) -> list[dict[str, str]]:
    """Format one tactic sheet."""
    headers = header_map(worksheet)
    last_row = last_value_row(worksheet)
    last_col = max(headers.values()) if headers else worksheet.max_column

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    set_column_widths(worksheet, TACTIC_WIDTHS)
    format_header_row(worksheet, HEADER_FILL)

    bool_columns = {headers[name] for name in BOOL_HEADERS if name in headers}
    center_columns = bool_columns | {headers[name] for name in ("Technical", "LLM") if name in headers}
    wrap_columns = {headers[name] for name in ("Title", "Command_match_rule", "Commandline_evasion", "Trigger rule") if name in headers}

    for row in range(2, last_row + 1):
        row_has_value = any(worksheet.cell(row, column).value not in (None, "") for column in range(1, last_col + 1))
        base_fill = ALT_ROW_FILL if row % 2 == 0 else WHITE_FILL

        for column in range(1, last_col + 1):
            cell = worksheet.cell(row, column)
            cell.border = THIN_BORDER
            cell.fill = base_fill
            cell.font = Font(name="Calibri", size=10, color="1F1F1F")
            if column in center_columns:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=column in wrap_columns)
            if column in bool_columns:
                format_bool_cell(cell)

        if row_has_value:
            texts = []
            for header in ("Title", "Command_match_rule", "Commandline_evasion", "Trigger rule"):
                column = headers.get(header)
                if column:
                    texts.append((cell_text(worksheet.cell(row, column).value), TACTIC_WIDTHS.get(header, 30)))
            worksheet.row_dimensions[row].height = row_height_for_texts(texts)
        else:
            worksheet.row_dimensions[row].height = 8

    return extend_tables_to_last_row(worksheet, last_row, last_col)


def format_summary_sheet(worksheet: Any) -> None:
    """Format Fill_Summary or summary-like sheets."""
    headers = header_map(worksheet)
    last_row = last_value_row(worksheet)
    last_col = max(headers.values()) if headers else worksheet.max_column

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    set_column_widths(worksheet, SUMMARY_WIDTHS)
    format_header_row(worksheet, SUMMARY_HEADER_FILL)

    for row in range(2, last_row + 1):
        values = [worksheet.cell(row, column).value for column in range(1, last_col + 1)]
        non_empty = [value for value in values if value not in (None, "")]
        is_section = len(non_empty) == 1 and worksheet.cell(row, 1).value not in (None, "")
        fill = SECTION_FILL if is_section else (ALT_ROW_FILL if row % 2 == 0 else WHITE_FILL)
        worksheet.row_dimensions[row].height = 22 if is_section else 24

        for column in range(1, last_col + 1):
            cell = worksheet.cell(row, column)
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10, bold=is_section, color="1F1F1F")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(description="Format final_result.xlsx for easier review.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--no-backup", action="store_true", help="Do not create a .format-backup-*.xlsx file.")
    return parser.parse_args()


def main() -> int:
    """Format the workbook in-place."""
    args = parse_args()
    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"workbook not found: {workbook_path}")

    backup_path: Path | None = None
    if not args.no_backup:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = workbook_path.with_name(f"{workbook_path.stem}.format-backup-{timestamp}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup_path)

    workbook = load_workbook(workbook_path)
    formatted_tactic_sheets = 0
    formatted_summary_sheets = 0
    table_updates: list[dict[str, str]] = []

    for worksheet in workbook.worksheets:
        headers = header_map(worksheet)
        if is_tactic_sheet(headers):
            table_updates.extend(format_tactic_sheet(worksheet))
            formatted_tactic_sheets += 1
        elif normalize_sheet_name(worksheet.title) in SUMMARY_SHEET_NAMES:
            format_summary_sheet(worksheet)
            formatted_summary_sheets += 1

    workbook.save(workbook_path)

    print(f"Formatted workbook: {workbook_path}")
    print(f"Tactic sheets formatted: {formatted_tactic_sheets}")
    print(f"Summary sheets formatted: {formatted_summary_sheets}")
    print(f"Table ranges updated: {len(table_updates)}")
    if backup_path:
        print(f"Backup: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
